"""
Tests for Gosreestr-style yearly report.
"""
from datetime import date
from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from openpyxl import load_workbook

from trials_app.models import (
    Application,
    ApplicationOblastState,
    Culture,
    GroupCulture,
    Oblast,
    Originator,
    SortOriginator,
    SortRecord,
)
from trials_app.services.reports import (
    build_applications_report_rows,
    build_applications_report_totals,
    build_applications_report_workbook,
)


class ReportServiceTest(TestCase):

    def setUp(self):
        self.group = GroupCulture.objects.create(group_culture_id=1, name='Зерновые культуры')
        self.culture = Culture.objects.create(culture_id=1, name='пшеница', group_culture=self.group)
        self.oblast1 = Oblast.objects.create(name='Костанайская', code='KOS')
        self.oblast2 = Oblast.objects.create(name='Акмолинская', code='AKM')
        self.user = User.objects.create_user(username='test', password='test')

        self.orig_domestic = Originator.objects.create(originator_id=1, name='ТОО НИИ', country='KZ', is_nanoc=True)
        self.orig_foreign = Originator.objects.create(originator_id=2, name='Bayer AG', country='DE')

    def _make_app(self, number, submission_date, sort_name, culture=None, originators=None, oblast_statuses=None):
        sr = SortRecord.objects.create(
            sort_id=SortRecord.objects.count() + 100,
            name=sort_name,
            culture=culture or self.culture,
        )
        if originators:
            for orig, pct in originators:
                SortOriginator.objects.create(sort_record=sr, originator=orig, percentage=pct)

        app = Application.objects.create(
            application_number=number,
            submission_date=submission_date,
            sort_record=sr,
            applicant='test',
            status='submitted',
            created_by=self.user,
        )

        if oblast_statuses:
            for oblast, status, year in oblast_statuses:
                ApplicationOblastState.objects.create(
                    application=app, oblast=oblast, status=status, decision_year=year,
                )

        return app

    def test_empty_queryset(self):
        rows = build_applications_report_rows(Application.objects.none())
        self.assertEqual(rows, [])

    def test_single_approved_app(self):
        self._make_app(
            'APP-1', date(2023, 5, 1), 'Сорт 1',
            originators=[(self.orig_domestic, 100)],
            oblast_statuses=[(self.oblast1, 'approved', 2023)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['year'], 2023)
        self.assertEqual(rows[0]['incoming']['total'], 1)
        self.assertEqual(rows[0]['incoming']['domestic'], 1)
        self.assertEqual(rows[0]['included']['total'], 1)
        self.assertEqual(rows[0]['removed']['total'], 0)
        self.assertEqual(rows[0]['continued']['total'], 0)

    def test_resolved_status_priority(self):
        """approved > continued > removed"""
        self._make_app(
            'APP-2', date(2023, 3, 1), 'Сорт 2',
            oblast_statuses=[
                (self.oblast1, 'approved', 2023),
                (self.oblast2, 'planned', None),
            ],
        )

        rows = build_applications_report_rows(Application.objects.all())
        # approved takes priority over planned (continued)
        self.assertEqual(rows[0]['included']['total'], 1)
        self.assertEqual(rows[0]['continued']['total'], 0)

    def test_continued_statuses(self):
        """planned, in_trial → continued"""
        self._make_app(
            'APP-3', date(2024, 6, 1), 'Сорт 3',
            oblast_statuses=[(self.oblast1, 'in_trial', 2024)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        self.assertEqual(rows[0]['continued']['total'], 1)
        self.assertEqual(rows[0]['included']['total'], 0)

    def test_removed_statuses(self):
        """removed → removed bucket"""
        self._make_app(
            'APP-4', date(2022, 1, 15), 'Сорт 4',
            oblast_statuses=[(self.oblast1, 'removed', 2022)],
        )
        self._make_app(
            'APP-5', date(2022, 4, 1), 'Сорт 5',
            oblast_statuses=[(self.oblast1, 'removed', 2022)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        self.assertEqual(rows[0]['removed']['total'], 2)

    def test_withdrawn_excluded(self):
        """withdrawn doesn't appear in any bucket"""
        self._make_app(
            'APP-6', date(2023, 1, 1), 'Сорт 6',
            oblast_statuses=[(self.oblast1, 'withdrawn', 2023)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        self.assertEqual(len(rows), 0)  # no resolved status

    def test_report_year_basic(self):
        """Non-winter crop → submission_date.year"""
        self._make_app(
            'APP-7', date(2024, 6, 15), 'Яровой сорт',
            oblast_statuses=[(self.oblast1, 'approved', 2024)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        self.assertEqual(rows[0]['year'], 2024)

    def test_origin_types(self):
        """foreign, domestic, joint detection"""
        # Foreign only
        self._make_app(
            'APP-F', date(2023, 1, 1), 'Foreign Sort',
            originators=[(self.orig_foreign, 100)],
            oblast_statuses=[(self.oblast1, 'approved', 2023)],
        )
        # Domestic only
        self._make_app(
            'APP-D', date(2023, 2, 1), 'Domestic Sort',
            originators=[(self.orig_domestic, 100)],
            oblast_statuses=[(self.oblast1, 'approved', 2023)],
        )
        # Joint
        self._make_app(
            'APP-J', date(2023, 3, 1), 'Joint Sort',
            originators=[(self.orig_foreign, 50), (self.orig_domestic, 50)],
            oblast_statuses=[(self.oblast1, 'approved', 2023)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        inc = rows[0]['incoming']
        self.assertEqual(inc['total'], 3)
        self.assertEqual(inc['foreign'], 1)
        self.assertEqual(inc['domestic'], 1)
        self.assertEqual(inc['joint'], 1)

    def test_empty_originators_default_domestic(self):
        """No originators → domestic"""
        self._make_app(
            'APP-NO', date(2023, 1, 1), 'No Orig Sort',
            oblast_statuses=[(self.oblast1, 'planned', None)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        self.assertEqual(rows[0]['incoming']['domestic'], 1)

    def test_nanoc_flag(self):
        self._make_app(
            'APP-N', date(2023, 1, 1), 'Nanoc Sort',
            originators=[(self.orig_domestic, 100)],  # is_nanoc=True
            oblast_statuses=[(self.oblast1, 'approved', 2023)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        self.assertEqual(rows[0]['incoming']['nanoc'], 1)
        self.assertEqual(rows[0]['included']['nanoc'], 1)

    def test_multiple_apps_same_year(self):
        """Multiple apps in same year aggregated correctly"""
        self._make_app('M1', date(2023, 1, 1), 'Sort M1', oblast_statuses=[(self.oblast1, 'approved', 2023)])
        self._make_app('M2', date(2023, 6, 1), 'Sort M2', oblast_statuses=[(self.oblast1, 'planned', None)])
        self._make_app('M3', date(2023, 9, 1), 'Sort M3', oblast_statuses=[(self.oblast1, 'removed', 2023)])

        rows = build_applications_report_rows(Application.objects.all())
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['incoming']['total'], 3)
        self.assertEqual(rows[0]['included']['total'], 1)
        self.assertEqual(rows[0]['continued']['total'], 1)
        self.assertEqual(rows[0]['removed']['total'], 1)

    def test_totals(self):
        self._make_app('T1', date(2022, 1, 1), 'S1', oblast_statuses=[(self.oblast1, 'approved', 2022)])
        self._make_app('T2', date(2023, 1, 1), 'S2', oblast_statuses=[(self.oblast1, 'removed', 2023)])

        rows = build_applications_report_rows(Application.objects.all())
        totals = build_applications_report_totals(rows)

        self.assertEqual(totals['incoming']['total'], 2)
        self.assertEqual(totals['included']['total'], 1)
        self.assertEqual(totals['removed']['total'], 1)

    def test_percent_calculation(self):
        self._make_app(
            'P1', date(2023, 1, 1), 'Pct Sort',
            originators=[(self.orig_domestic, 100)],  # nanoc=True
            oblast_statuses=[(self.oblast1, 'approved', 2023)],
        )

        rows = build_applications_report_rows(Application.objects.all())
        # incoming: 1 total, 1 nanoc → 100%
        self.assertAlmostEqual(rows[0]['incoming']['percent'], 1.0)

    def test_workbook_structure(self):
        self._make_app('W1', date(2023, 1, 1), 'WB Sort', oblast_statuses=[(self.oblast1, 'approved', 2023)])

        rows = build_applications_report_rows(Application.objects.all())
        totals = build_applications_report_totals(rows)
        content = build_applications_report_workbook(rows, totals)

        wb = load_workbook(BytesIO(content))
        ws = wb.active

        self.assertEqual(ws.title, 'Отчёт')
        self.assertEqual(ws.max_column, 25)
        self.assertEqual(ws.cell(1, 1).value, 'Год')
        self.assertEqual(ws.cell(1, 2).value, 'Поступило')
        # Data row + totals row + 2 header rows
        self.assertEqual(ws.max_row, 4)  # 2 headers + 1 data + 1 totals
        self.assertEqual(ws.cell(ws.max_row, 1).value, 'Итого')

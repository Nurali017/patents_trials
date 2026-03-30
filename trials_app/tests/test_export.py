"""
Tests for Excel export of applications.
"""
from datetime import date
from io import BytesIO

from django.test import TestCase, override_settings
from openpyxl import load_workbook
from rest_framework.test import APIClient

from trials_app.models import (
    Application,
    ApplicationOblastState,
    Culture,
    GroupCulture,
    Oblast,
    Originator,
    SortOriginator,
    SortRecord,
)
from trials_app.services.exporters import build_applications_export_workbook


class ExportWorkbookTest(TestCase):
    """Test the exporter service directly."""

    def setUp(self):
        self.group = GroupCulture.objects.create(group_culture_id=1, name='Зерновые культуры')
        self.culture = Culture.objects.create(culture_id=1, name='пшеница', group_culture=self.group)
        self.sort = SortRecord.objects.create(sort_id=1, name='Тестовый сорт', public_code='TEST-1', culture=self.culture)
        self.oblast1 = Oblast.objects.create(name='Костанайская', code='KOS')
        self.oblast2 = Oblast.objects.create(name='Акмолинская', code='AKM')
        self.oblast3 = Oblast.objects.create(name='Павлодарская', code='PAV')

        from django.contrib.auth.models import User
        self.user = User.objects.create_user(username='test', password='test')

        self.app = Application.objects.create(
            application_number='TEST-001',
            submission_date=date(2024, 1, 15),
            sort_record=self.sort,
            applicant='Тестовый заявитель',
            status='submitted',
            created_by=self.user,
        )
        self.app.target_oblasts.add(self.oblast1, self.oblast2, self.oblast3)

    def test_export_returns_xlsx(self):
        """Export produces valid xlsx bytes."""
        content = build_applications_export_workbook(Application.objects.all())
        self.assertTrue(len(content) > 0)

        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.title, 'Заявки')
        self.assertEqual(ws[1][0].value, 'Номер заявки')
        self.assertEqual(ws.max_column, 13)

    def test_one_app_multiple_oblast_states(self):
        """One application with 3 oblast states produces 3 rows."""
        ApplicationOblastState.objects.create(application=self.app, oblast=self.oblast1, status='approved', decision_year=2024)
        ApplicationOblastState.objects.create(application=self.app, oblast=self.oblast2, status='planned')
        ApplicationOblastState.objects.create(application=self.app, oblast=self.oblast3, status='removed', decision_year=2023)

        content = build_applications_export_workbook(Application.objects.all())
        wb = load_workbook(BytesIO(content))
        ws = wb.active

        # 1 header + 3 data rows
        self.assertEqual(ws.max_row, 4)

        # Check oblast names in rows
        oblasts = {ws.cell(row=r, column=10).value for r in range(2, 5)}
        self.assertEqual(oblasts, {'Костанайская', 'Акмолинская', 'Павлодарская'})

    def test_empty_originators_no_crash(self):
        """Export doesn't crash when sort has no originators."""
        content = build_applications_export_workbook(Application.objects.all())
        wb = load_workbook(BytesIO(content))
        ws = wb.active

        # Should have at least header + 1 row (app without oblast states)
        self.assertGreaterEqual(ws.max_row, 2)

        # Origin type should be '-', NANOC 'Нет', originators empty
        row2 = [ws.cell(row=2, column=c).value for c in range(1, 14)]
        self.assertEqual(row2[6], '-')       # Происхождение
        self.assertEqual(row2[7], 'Нет')     # НАНОЦ
        self.assertIn(row2[8], ('', None))  # Оригинаторы (openpyxl reads empty cell as None)

    def test_with_originators(self):
        """Export correctly shows originator data."""
        orig = Originator.objects.create(originator_id=1, name='ТОО Тестовый НИИ', country='KZ', is_nanoc=True)
        SortOriginator.objects.create(sort_record=self.sort, originator=orig, percentage=100)

        ApplicationOblastState.objects.create(application=self.app, oblast=self.oblast1, status='approved')

        content = build_applications_export_workbook(Application.objects.all())
        wb = load_workbook(BytesIO(content))
        ws = wb.active

        row2 = [ws.cell(row=2, column=c).value for c in range(1, 14)]
        self.assertEqual(row2[6], 'Отечественный')
        self.assertEqual(row2[7], 'Да')
        self.assertEqual(row2[8], 'ТОО Тестовый НИИ')

    def test_oblast_filter(self):
        """filter_oblast_id restricts exported states."""
        ApplicationOblastState.objects.create(application=self.app, oblast=self.oblast1, status='approved')
        ApplicationOblastState.objects.create(application=self.app, oblast=self.oblast2, status='planned')

        content = build_applications_export_workbook(
            Application.objects.all(),
            filter_oblast_id=self.oblast1.id,
        )
        wb = load_workbook(BytesIO(content))
        ws = wb.active

        self.assertEqual(ws.max_row, 2)  # header + 1 row (only Костанайская)
        self.assertEqual(ws.cell(row=2, column=10).value, 'Костанайская')

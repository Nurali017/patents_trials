"""
Report service for Gosreestr-style yearly report.

Builds application snapshots with:
- report_year (winter crop + Jan → same year, else +1)
- resolved_status (priority: approved > continued > removed)
- origin_type (foreign/joint/domestic from SortOriginator)
- is_nanoc

Aggregates into 4 buckets per year: incoming, included, removed, continued.
"""
import io
from collections import defaultdict
from dataclasses import dataclass, field

from django.db.models import Prefetch
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Statuses that count as "continued" (ongoing testing)
CONTINUED_STATUSES = {'in_trial', 'planned'}

# Statuses that count as "removed"
REMOVED_STATUSES = {'removed'}


@dataclass
class ReportBucket:
    total: int = 0
    foreign: int = 0
    joint: int = 0
    domestic: int = 0
    nanoc: int = 0
    percent: float = None

    def as_dict(self):
        return {
            'total': self.total,
            'foreign': self.foreign,
            'joint': self.joint,
            'domestic': self.domestic,
            'nanoc': self.nanoc,
            'percent': self.percent,
        }


def _compute_report_year(submission_date, is_winter):
    """Compute report year using legacy rule."""
    if not submission_date:
        return 2020
    year = submission_date.year
    if is_winter and submission_date.month == 1:
        return year
    return year


def _compute_origin_type(sort_originators):
    """Compute origin type. Empty → domestic (legacy rule)."""
    if not sort_originators:
        return 'domestic'

    has_foreign = any(so.originator.is_foreign for so in sort_originators)
    has_domestic = any(not so.originator.is_foreign for so in sort_originators)

    if has_foreign and has_domestic:
        return 'joint'
    if has_foreign:
        return 'foreign'
    return 'domestic'


def _compute_is_nanoc(sort_originators):
    """Check NANOC. Empty → False."""
    if not sort_originators:
        return False
    return any(so.originator.is_nanoc for so in sort_originators)


def _resolve_status(oblast_statuses):
    """
    Resolve application status from all oblast states.
    Priority: approved > continued > removed.
    Withdrawn excluded.
    """
    statuses = {s for s in oblast_statuses if s != 'withdrawn'}
    if not statuses:
        return None

    if 'approved' in statuses:
        return 'approved'

    if statuses & CONTINUED_STATUSES:
        return 'continued'

    if statuses & REMOVED_STATUSES:
        return 'removed'

    return None


def build_applications_report_rows(queryset):
    """
    Build report rows from Application queryset.

    Returns list of dicts with year + 4 buckets.
    """
    from trials_app.models import ApplicationOblastState, SortOriginator

    applications = queryset.select_related(
        'sort_record__culture__group_culture',
    ).prefetch_related(
        Prefetch(
            'oblast_states',
            queryset=ApplicationOblastState.objects.filter(is_deleted=False),
            to_attr='_report_oblast_states',
        ),
        Prefetch(
            'sort_record__sort_originators',
            queryset=SortOriginator.objects.select_related('originator'),
            to_attr='_report_sort_originators',
        ),
    )

    # Build snapshots per application
    year_data = defaultdict(lambda: {
        'incoming': ReportBucket(),
        'included': ReportBucket(),
        'removed': ReportBucket(),
        'continued': ReportBucket(),
    })

    seen_apps = set()

    for app in applications:
        # Dedup by application_number
        if app.application_number in seen_apps:
            continue
        seen_apps.add(app.application_number)

        sr = app.sort_record
        if not sr:
            continue

        # Compute snapshot fields
        is_winter = getattr(sr.culture, 'is_winter', False) if sr.culture else False
        report_year = _compute_report_year(app.submission_date, is_winter)

        sort_origs = getattr(sr, '_report_sort_originators', []) or []
        origin = _compute_origin_type(sort_origs)
        nanoc = _compute_is_nanoc(sort_origs)

        # Get all oblast statuses
        oblast_states = getattr(app, '_report_oblast_states', []) or []
        all_statuses = [s.status for s in oblast_states]
        resolved = _resolve_status(all_statuses)

        if resolved is None:
            continue

        buckets = year_data[report_year]

        # Incoming: all with any resolved status
        _add_to_bucket(buckets['incoming'], origin, nanoc)

        # Categorize
        if resolved == 'approved':
            _add_to_bucket(buckets['included'], origin, nanoc)
        elif resolved == 'removed':
            _add_to_bucket(buckets['removed'], origin, nanoc)
        elif resolved == 'continued':
            _add_to_bucket(buckets['continued'], origin, nanoc)

    # Build rows sorted by year
    rows = []
    for year in sorted(year_data.keys()):
        buckets = year_data[year]
        incoming_nanoc = buckets['incoming'].nanoc

        # Calculate percentages (legacy formula)
        buckets['incoming'].percent = _safe_pct(buckets['incoming'].nanoc, buckets['incoming'].total)
        buckets['included'].percent = _safe_pct(buckets['included'].nanoc, incoming_nanoc)
        buckets['removed'].percent = _safe_pct(buckets['removed'].nanoc, incoming_nanoc)
        buckets['continued'].percent = _safe_pct(buckets['continued'].nanoc, incoming_nanoc)

        rows.append({
            'year': year,
            'incoming': buckets['incoming'].as_dict(),
            'included': buckets['included'].as_dict(),
            'removed': buckets['removed'].as_dict(),
            'continued': buckets['continued'].as_dict(),
        })

    return rows


def build_applications_report_totals(rows):
    """Sum all rows into totals."""
    totals = {
        'incoming': ReportBucket(),
        'included': ReportBucket(),
        'removed': ReportBucket(),
        'continued': ReportBucket(),
    }

    for row in rows:
        for key in ('incoming', 'included', 'removed', 'continued'):
            b = totals[key]
            r = row[key]
            b.total += r['total']
            b.foreign += r['foreign']
            b.joint += r['joint']
            b.domestic += r['domestic']
            b.nanoc += r['nanoc']

    incoming_nanoc = totals['incoming'].nanoc
    totals['incoming'].percent = _safe_pct(totals['incoming'].nanoc, totals['incoming'].total)
    totals['included'].percent = _safe_pct(totals['included'].nanoc, incoming_nanoc)
    totals['removed'].percent = _safe_pct(totals['removed'].nanoc, incoming_nanoc)
    totals['continued'].percent = _safe_pct(totals['continued'].nanoc, incoming_nanoc)

    return {k: v.as_dict() for k, v in totals.items()}


def build_applications_report_workbook(rows, totals):
    """Build Excel workbook for the report."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Отчёт'

    # Two-level header
    sections = ['Поступило', 'Включено в госреестр', 'Снято с испытаний', 'Продолжены испытания']
    sub_headers = ['Всего', 'Заруб.', 'Совм.', 'Отеч.', 'НАНОЦ', '%']

    # Row 1: merged section headers
    ws.cell(row=1, column=1, value='Год')
    col = 2
    for section in sections:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 5)
        ws.cell(row=1, column=col, value=section)
        col += 6

    # Row 2: sub-headers
    ws.cell(row=2, column=1, value='')
    col = 2
    for _ in sections:
        for sh in sub_headers:
            ws.cell(row=2, column=col, value=sh)
            col += 1

    ws.freeze_panes = 'A3'

    # Data rows
    bucket_keys = ['incoming', 'included', 'removed', 'continued']
    for row_data in rows:
        row_values = [row_data['year']]
        for key in bucket_keys:
            b = row_data[key]
            pct = f'{b["percent"]:.1%}' if b['percent'] is not None else '-'
            row_values.extend([b['total'], b['foreign'], b['joint'], b['domestic'], b['nanoc'], pct])
        ws.append(row_values)

    # Totals row
    totals_values = ['Итого']
    for key in bucket_keys:
        b = totals[key]
        pct = f'{b["percent"]:.1%}' if b['percent'] is not None else '-'
        totals_values.extend([b['total'], b['foreign'], b['joint'], b['domestic'], b['nanoc'], pct])
    ws.append(totals_values)

    # Auto-width
    for col_idx in range(1, 26):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    ws.column_dimensions['A'].width = 6

    # Autofilter
    if ws.max_row > 2:
        ws.auto_filter.ref = f'A2:{get_column_letter(25)}{ws.max_row}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _add_to_bucket(bucket, origin, nanoc):
    """Add one application to a bucket."""
    bucket.total += 1
    if origin == 'foreign':
        bucket.foreign += 1
    elif origin == 'joint':
        bucket.joint += 1
    else:
        bucket.domestic += 1
    if nanoc:
        bucket.nanoc += 1


def _safe_pct(numerator, denominator):
    """Safe percentage calculation. Returns None if division impossible."""
    if not denominator:
        return None
    return numerator / denominator

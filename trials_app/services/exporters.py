"""
Excel export for Applications with oblast states.

One row per ApplicationOblastState.
13 columns as specified in migration plan.
"""
import io
from datetime import date

from django.db.models import Prefetch
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


HEADERS = [
    'Номер заявки',
    'Дата подачи',
    'Группа культур',
    'Культура',
    'Сорт',
    'Селекционный номер',
    'Происхождение',
    'НАНОЦ',
    'Оригинаторы',
    'Область',
    'Статус области',
    'Год решения',
    'Статус заявки',
]

STATUS_DISPLAY = {
    'planned': 'Запланировано',
    'trial_plan_created': 'План создан',
    'trial_created': 'Испытание создано',
    'trial_in_progress': 'Испытание идет',
    'trial_completed': 'Испытание завершено',
    'decision_pending': 'Ожидает решения',
    'decision_made': 'Решение принято',
    'approved': 'Одобрено',
    'rejected': 'Отклонено',
    'continue': 'Продолжить',
    'removed': 'Снят',
    'withdrawn': 'Отозван',
    'submitted': 'Подана',
    'distributed': 'Распределена',
    'in_progress': 'Испытания проводятся',
    'completed': 'Завершена',
    'registered': 'В реестре',
    'draft': 'Черновик',
}


def _compute_origin_type(sort_originators):
    """Compute origin type from sort originators (legacy Gosreestr logic)."""
    if not sort_originators:
        return '-'

    has_foreign = any(so.originator.is_foreign for so in sort_originators)
    has_domestic = any(not so.originator.is_foreign for so in sort_originators)

    if has_foreign and has_domestic:
        return 'Совместный'
    if has_foreign:
        return 'Зарубежный'
    return 'Отечественный'


def _compute_is_nanoc(sort_originators):
    """Check if any originator is NANOC."""
    if not sort_originators:
        return 'Нет'
    return 'Да' if any(so.originator.is_nanoc for so in sort_originators) else 'Нет'


def _format_originators(sort_originators):
    """Format originators as display string."""
    if not sort_originators:
        return ''  # noqa: will be written as empty cell

    sorted_origs = sorted(sort_originators, key=lambda so: (-so.percentage, so.originator.name))
    parts = []
    for so in sorted_origs:
        if so.percentage == 100 and len(sorted_origs) == 1:
            parts.append(so.originator.name)
        else:
            parts.append(f'{so.originator.name} ({so.percentage}%)')
    return '; '.join(parts)


def build_applications_export_workbook(queryset, filter_oblast_id=None):
    """
    Build Excel workbook from Application queryset.

    One row per ApplicationOblastState.
    Returns bytes (xlsx content).

    Args:
        queryset: Application queryset (already filtered by ApplicationFilter)
        filter_oblast_id: If set, export only this oblast's states (for strict oblast filtering)
    """
    from trials_app.models import ApplicationOblastState, SortOriginator

    wb = Workbook()
    ws = wb.active
    ws.title = 'Заявки'

    # Header
    ws.append(HEADERS)
    ws.freeze_panes = 'A2'

    # Track column widths
    col_widths = [len(h) for h in HEADERS]

    # Optimized prefetch — read from cache, no extra queries in loop
    applications = queryset.select_related(
        'sort_record__culture__group_culture',
    ).prefetch_related(
        Prefetch(
            'oblast_states',
            queryset=ApplicationOblastState.objects.filter(is_deleted=False).select_related('oblast'),
            to_attr='_prefetched_oblast_states',
        ),
        Prefetch(
            'sort_record__sort_originators',
            queryset=SortOriginator.objects.select_related('originator'),
            to_attr='_prefetched_sort_originators',
        ),
    )

    for app in applications:
        sr = app.sort_record
        if not sr:
            continue

        # Read from prefetched cache — no extra DB queries
        sort_origs = getattr(sr, '_prefetched_sort_originators', []) or []
        origin_type = _compute_origin_type(sort_origs)
        is_nanoc = _compute_is_nanoc(sort_origs)
        originators_display = _format_originators(sort_origs)

        # Culture info
        culture = sr.culture if sr else None
        group_name = culture.group_culture.name if culture and culture.group_culture else ''
        culture_name = culture.name if culture else ''

        # Read from prefetched cache
        oblast_states = getattr(app, '_prefetched_oblast_states', []) or []

        # Apply oblast filter if specified
        if filter_oblast_id is not None:
            oblast_states = [s for s in oblast_states if s.oblast_id == filter_oblast_id]

        if not oblast_states:
            # Application without oblast states — still export one row
            row = [
                app.application_number,
                app.submission_date,
                group_name,
                culture_name,
                sr.name if sr else '',
                (sr.public_code or '') if sr else '',
                origin_type,
                is_nanoc,
                originators_display,
                '',
                '',
                None,
                STATUS_DISPLAY.get(app.status, app.status),
            ]
            ws.append(row)
            _update_widths(col_widths, row)
        else:
            for state in oblast_states:
                row = [
                    app.application_number,
                    app.submission_date,
                    group_name,
                    culture_name,
                    sr.name if sr else '',
                    (sr.public_code or '') if sr else '',
                    origin_type,
                    is_nanoc,
                    originators_display,
                    state.oblast.name if state.oblast else '',
                    STATUS_DISPLAY.get(state.status, state.status),
                    state.decision_year,
                    STATUS_DISPLAY.get(app.status, app.status),
                ]
                ws.append(row)
                _update_widths(col_widths, row)

    # Apply column widths
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)

    # Autofilter
    if ws.max_row > 1:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{ws.max_row}'

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _update_widths(col_widths, row):
    """Update column width tracking."""
    for i, val in enumerate(row):
        if val is not None:
            col_widths[i] = max(col_widths[i], len(str(val)))

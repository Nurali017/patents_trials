from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils import timezone
from django.db import models as django_models
from .models import (
    Oblast, Region, ClimateZone, Indicator, GroupCulture, Culture, Originator, SortRecord, 
    Application, PlannedDistribution, TrialType, Trial, TrialParticipant, TrialResult, 
    TrialLaboratoryResult, Document, TrialPlan, TrialPlanParticipant, TrialPlanTrial, TrialPlanCulture,
    TrialPlanCultureTrialType, AnnualDecisionTable, AnnualDecisionItem
)
from .serializers import (
    OblastSerializer, RegionSerializer, ClimateZoneSerializer, IndicatorSerializer,
    GroupCultureSerializer, CultureSerializer, OriginatorSerializer, SortRecordSerializer, 
    ApplicationSerializer, TrialTypeSerializer, TrialSerializer, TrialParticipantSerializer,
    TrialResultSerializer, TrialLaboratoryResultSerializer, DocumentSerializer,
    TrialPlanSerializer, TrialPlanWriteSerializer, TrialPlanAddParticipantsSerializer,
    TrialPlanCultureSerializer, TrialPlanAddCultureSerializer,
    create_basic_trial_results, create_quality_trial_results
)
from .annual_decision_serializers import (
    AnnualDecisionTableSerializer, AnnualDecisionTableDetailSerializer, AnnualDecisionTableCreateSerializer,
    AnnualDecisionItemSerializer, AnnualDecisionItemDetailSerializer, AnnualDecisionItemUpdateSerializer
)
from .patents_integration import patents_api

class OblastViewSet(viewsets.ModelViewSet):
    """Области для испытаний"""
    queryset = Oblast.objects.filter(is_deleted=False)
    serializer_class = OblastSerializer
    pagination_class = None  # Отключаем пагинацию - возвращаем все области
    
    def get_permissions(self):
        """Чтение - всем, изменение - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

class ClimateZoneViewSet(viewsets.ModelViewSet):
    """Природно-климатические зоны"""
    queryset = ClimateZone.objects.filter(is_deleted=False)
    serializer_class = ClimateZoneSerializer
    pagination_class = None  # Отключаем пагинацию - возвращаем все зоны
    
    def get_permissions(self):
        """Чтение - всем, изменение - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

class RegionViewSet(viewsets.ModelViewSet):
    """Сортоиспытательные участки (ГСУ)"""
    queryset = Region.objects.filter(is_deleted=False).select_related('oblast', 'climate_zone')
    serializer_class = RegionSerializer
    pagination_class = None  # Отключаем пагинацию - возвращаем все регионы
    
    def get_permissions(self):
        """Чтение - всем, изменение - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]
    
    def get_queryset(self):
        """Фильтрация по области"""
        queryset = super().get_queryset()
        
        # Фильтр по области
        oblast_id = self.request.query_params.get('oblast')
        if oblast_id:
            queryset = queryset.filter(oblast_id=oblast_id)
        
        return queryset

class IndicatorViewSet(viewsets.ModelViewSet):
    """
    Показатели измерений
    
    Фильтрация:
        ?culture_id=<id> - показатели для конкретной культуры (через группу культуры)
        ?group_culture_id=<id> - показатели для группы культур
        ?culture_group=<id> - алиас для group_culture_id
        ?is_universal=true - универсальные показатели
        ?category=common|quality|specific - по категории
        ?is_quality=true|false - лабораторные (true) или основные (false) показатели
        ?search=<text> - поиск по названию или коду
    """
    queryset = Indicator.objects.filter(is_deleted=False).prefetch_related('group_cultures')
    serializer_class = IndicatorSerializer
    
    def get_permissions(self):
        """Чтение - всем, изменение - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]
    
    def get_queryset(self):
        """Фильтрация по культуре или группе культур (новая логика через group_cultures)"""
        queryset = super().get_queryset()
        
        # Фильтр по культуре - получаем показатели группы этой культуры
        culture_id = self.request.query_params.get('culture_id')
        if culture_id:
            try:
                culture = Culture.objects.get(id=culture_id, is_deleted=False)
                if culture.group_culture:
                    # Показатели группы этой культуры (НОВАЯ ЛОГИКА)
                    queryset = queryset.filter(
                        group_cultures=culture.group_culture
                    ).distinct()
                else:
                    # Если у культуры нет группы - вернуть пустой набор
                    queryset = queryset.none()
            except Culture.DoesNotExist:
                queryset = queryset.none()
        
        # Фильтр по группе культур напрямую (два варианта: group_culture_id и culture_group)
        group_culture_id = self.request.query_params.get('group_culture_id') or self.request.query_params.get('culture_group')
        if group_culture_id:
            queryset = queryset.filter(
                group_cultures__id=group_culture_id
            ).distinct()
        
        # Фильтр по категории
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        # Фильтр по is_quality (лабораторные/основные показатели)
        is_quality = self.request.query_params.get('is_quality')
        if is_quality is not None:
            if is_quality.lower() == 'true':
                queryset = queryset.filter(is_quality=True)
            elif is_quality.lower() == 'false':
                queryset = queryset.filter(is_quality=False)
        
        # Фильтр только универсальные
        is_universal = self.request.query_params.get('is_universal')
        if is_universal == 'true':
            queryset = queryset.filter(is_universal=True)
        
        # Поиск по названию или коду
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                django_models.Q(name__icontains=search) |
                django_models.Q(code__icontains=search)
            )
        
        return queryset

class TrialTypeViewSet(viewsets.ModelViewSet):
    """Типы испытаний (КСИ, ООС, ДЮС-ТЕСТ и т.д.)"""
    queryset = TrialType.objects.filter(is_deleted=False)
    serializer_class = TrialTypeSerializer
    pagination_class = None  # Отключаем пагинацию - возвращаем все типы испытаний
    
    def get_permissions(self):
        """Чтение - всем, изменение - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]


class GroupCultureViewSet(viewsets.ModelViewSet):
    """
    Группы культур (Зерновые, Овощные и т.д.)
    
    Локальная копия из Patents Service для автономной работы
    """
    queryset = GroupCulture.objects.filter(is_deleted=False)
    serializer_class = GroupCultureSerializer
    
    def get_permissions(self):
        """Чтение - всем, изменение/синхронизация - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]


class CultureViewSet(viewsets.ModelViewSet):
    """
    Культуры растений (Пшеница яровая, Ячмень и т.д.)
    
    Локальная копия из Patents Service для автономной работы
    """
    queryset = Culture.objects.filter(is_deleted=False)
    serializer_class = CultureSerializer
    
    def get_permissions(self):
        """Чтение - всем, изменение/синхронизация - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]
    
    @action(detail=True, methods=['post'], url_path='sync')
    def sync(self, request, pk=None):
        """
        Синхронизировать культуру с Patents Service
        
        POST /api/v1/cultures/{id}/sync/
        """
        culture = self.get_object()
        if culture.sync_from_patents():
            return Response({
                'success': True,
                'message': f'Культура {culture.name} успешно синхронизирована'
            })
        else:
            return Response({
                'success': False,
                'error': 'Failed to sync with Patents Service'
            }, status=500)


class OriginatorViewSet(viewsets.ModelViewSet):
    """
    Оригинаторы (создатели сортов)
    
    Локальная копия данных из Patents Service
    """
    queryset = Originator.objects.filter(is_deleted=False)
    serializer_class = OriginatorSerializer
    
    def get_permissions(self):
        """Чтение - всем, изменение/синхронизация - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]
    
    @action(detail=True, methods=['post'], url_path='sync')
    def sync(self, request, pk=None):
        """
        Синхронизировать оригинатора с Patents Service
        
        POST /api/v1/originators/{id}/sync/
        """
        originator = self.get_object()
        if originator.sync_from_patents():
            return Response({
                'success': True,
                'message': f'Оригинатор {originator.name} успешно синхронизирован'
            })
        else:
            return Response({
                'success': False,
                'error': 'Failed to sync with Patents Service'
            }, status=500)

class SortRecordViewSet(viewsets.ModelViewSet):
    """
    Записи о сортах для испытаний
    
    Хранит ВСЕ данные сорта локально для автономной работы
    Синхронизируется с Patents Service
    """
    queryset = SortRecord.objects.filter(is_deleted=False)
    serializer_class = SortRecordSerializer
    
    def get_permissions(self):
        """Чтение - всем, изменение/синхронизация - только авторизованным"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]
    
    @action(detail=True, methods=['post'], url_path='sync')
    def sync(self, request, pk=None):
        """
        Синхронизировать сорт с Patents Service
        
        Обновляет все данные включая оригинаторов
        
        POST /api/v1/sort-records/{id}/sync/
        """
        sort_record = self.get_object()
        if sort_record.sync_from_patents():
            return Response({
                'success': True,
                'message': f'Сорт {sort_record.name} успешно синхронизирован',
                'data': SortRecordSerializer(sort_record).data
            })
        else:
            return Response({
                'success': False,
                'error': 'Failed to sync with Patents Service'
            }, status=500)
    
    @action(detail=False, methods=['post'], url_path='sync-all')
    def sync_all(self, request):
        """
        Синхронизировать все сорта с Patents Service
        
        POST /api/v1/sort-records/sync-all/
        """
        count = 0
        errors = []
        
        for sort_record in self.get_queryset():
            if sort_record.sync_from_patents():
                count += 1
            else:
                errors.append(sort_record.sort_id)
        
        return Response({
            'synced': count,
            'failed': len(errors),
            'failed_ids': errors
        })


class ApplicationViewSet(viewsets.ModelViewSet):
    """
    Управление заявками на сортоиспытания
    
    Заявка → Распределение по областям → Испытания → Решения → Реестр
    """
    queryset = Application.objects.filter(is_deleted=False)
    serializer_class = ApplicationSerializer
    permission_classes = [permissions.IsAuthenticated]  # Требуется авторизация
    
    def perform_create(self, serializer):
        """При создании заявки устанавливаем created_by"""
        serializer.save(created_by=self.request.user if self.request.user.is_authenticated else None)
    
    @action(detail=True, methods=['post'], url_path='distribute')
    def distribute(self, request, pk=None):
        """
        Распределить заявку по ГСУ
        
        Сохраняет параметры распределения в Application.
        Trial создается ВРУЧНУЮ сортопытом позже!
        
        POST /api/v1/applications/{id}/distribute/
        Body: {
            "distributions": [
                {
                    "region": 1,                    # ID ГСУ (обязательно)
                    "trial_type": "competitive",    # Код типа испытания
                    "planting_season": "spring"     # Опционально
                }
            ]
        }
        """
        application = self.get_object()
        
        if application.status not in ['draft', 'submitted']:
            return Response({
                'error': 'Application already distributed'
            }, status=400)
        
        distributions = request.data.get('distributions', [])
        if not distributions:
            return Response({
                'error': 'No distributions provided'
            }, status=400)
        
        # Валидация параметров
        validated_distributions = []
        for dist in distributions:
            # Проверка ГСУ
            try:
                region = Region.objects.get(id=dist['region'])
            except Region.DoesNotExist:
                return Response({
                    'error': f'Region with ID {dist["region"]} not found'
                }, status=400)
            
            # ПРОВЕРКА: ГСУ должен быть в одной из целевых областей
            target_oblast_ids = application.target_oblasts.values_list('id', flat=True)
            if region.oblast_id not in target_oblast_ids:
                return Response({
                    'error': f'Region "{region.name}" is not in target oblasts. '
                            f'Application targets: {", ".join(application.target_oblasts.values_list("name", flat=True))}'
                }, status=400)
            
            # Проверка типа испытания
            trial_type_obj = None
            if 'trial_type' in dist:
                try:
                    trial_type_obj = TrialType.objects.get(code=dist['trial_type'])
                except TrialType.DoesNotExist:
                    return Response({
                        'error': f'Trial type "{dist["trial_type"]}" not found'
                    }, status=400)
            
            validated_distributions.append({
                'region': region,
                'trial_type': trial_type_obj,
                'planting_season': dist.get('planting_season'),
                'raw_data': dist  # Сохраняем оригинальные данные для JSON
            })
        
        # Удаляем старые распределения (если переделывают)
        PlannedDistribution.objects.filter(application=application).delete()
        
        # Создаем записи в таблице PlannedDistribution
        created_distributions = []
        for validated in validated_distributions:
            planned_dist = PlannedDistribution.objects.create(
                application=application,
                region=validated['region'],
                trial_type=validated['trial_type'],
                planting_season=validated['planting_season'],
                created_by=request.user
            )
            created_distributions.append(planned_dist)
        
        # Обновляем статус заявки
        application.status = 'distributed'
        application.save()
        
        return Response({
            'success': True,
            'message': f'Application distributed with {len(distributions)} planned trial(s)',
            'planned_distributions': distributions,
            'application': ApplicationSerializer(application).data
        })
    
    @action(detail=True, methods=['get'], url_path='regional-trials')
    def regional_trials(self, request, pk=None):
        """
        Получить все испытания для заявки
        
        GET /api/v1/applications/{id}/regional-trials/
        
        Находит испытания через TrialParticipant, где application=данная заявка
        """
        application = self.get_object()
        
        # Найти все уникальные испытания через участников этой заявки
        from trials_app.models import TrialParticipant
        trial_ids = TrialParticipant.objects.filter(
            application=application,
            is_deleted=False
        ).values_list('trial_id', flat=True).distinct()
        
        trials = Trial.objects.filter(
            id__in=trial_ids,
            is_deleted=False
        )
        
        serializer = TrialSerializer(trials, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], url_path='decisions')
    def decisions(self, request, pk=None):
        """
        Получить все решения по испытаниям заявки
        
        GET /api/v1/applications/{id}/decisions/
        """
        application = self.get_object()
        
        # Найти все уникальные испытания через участников этой заявки
        from trials_app.models import TrialParticipant
        trial_ids = TrialParticipant.objects.filter(
            application=application,
            is_deleted=False
        ).values_list('trial_id', flat=True).distinct()
        
        trials = Trial.objects.filter(
            id__in=trial_ids,
            is_deleted=False
        ).exclude(
            decision__isnull=True
        ).exclude(decision='')
        
        decisions_data = []
        for trial in trials:
            decisions_data.append({
                'trial_id': trial.id,
                'region': trial.region.name,
                'oblast': trial.region.oblast.name,
                'decision': trial.decision,
                'decision_display': trial.get_decision_display(),
                'justification': trial.decision_justification,
                'recommendations': trial.decision_recommendations,
                'decision_date': trial.decision_date,
                'decided_by': trial.decided_by.username if trial.decided_by else None,
            })
        
        return Response(decisions_data)
    
    @action(detail=False, methods=['get'], url_path='statistics')
    def statistics(self, request):
        """
        Статистика по заявкам
        
        GET /api/v1/applications/statistics/
        """
        applications = self.get_queryset()
        
        return Response({
            'total': applications.count(),
            'by_status': {
                'draft': applications.filter(status='draft').count(),
                'submitted': applications.filter(status='submitted').count(),
                'distributed': applications.filter(status='distributed').count(),
                'in_progress': applications.filter(status='in_progress').count(),
                'completed': applications.filter(status='completed').count(),
                'registered': applications.filter(status='registered').count(),
                'rejected': applications.filter(status='rejected').count(),
            },
            'current_year': applications.filter(
                submission_date__year=timezone.now().year
            ).count(),
        })
    
    @action(detail=False, methods=['get'], url_path='cultures-for-region')
    def cultures_for_region(self, request):
        """
        Получить список культур с заявками для ГСУ
        
        Показывает какие культуры доступны в выбранном ГСУ
        с количеством заявок по каждой культуре.
        
        GET /api/v1/applications/cultures-for-region/?region_id=5
        
        Response:
        {
            "region_id": 5,
            "region_name": "Алматинский ГСУ",
            "oblast_id": 2,
            "oblast_name": "Алматинская область",
            "cultures": [
                {
                    "culture_id": 10,
                    "culture_name": "Пшеница яровая",
                    "applications_count": 3,
                    "pending_count": 3,
                    "in_trial_count": 1,
                    "sample_applications": ["APP-2025-001", "APP-2025-005"]
                }
            ]
        }
        """
        region_id = request.query_params.get('region_id')
        
        if not region_id:
            return Response({
                'error': 'region_id is required'
            }, status=400)
        
        try:
            region = Region.objects.get(id=region_id, is_deleted=False)
        except Region.DoesNotExist:
            return Response({'error': 'Region not found'}, status=404)
        
        # Найти все распределенные заявки для этого региона
        # Используем PlannedDistribution для точности
        planned_distributions = PlannedDistribution.objects.filter(
            region=region,
            status='planned',
            is_deleted=False
        ).select_related('application__sort_record__culture')
        
        # Сгруппировать по культурам
        cultures_data = {}
        for dist in planned_distributions:
            app = dist.application
            if not app or not app.sort_record or not app.sort_record.culture:
                continue
            
            culture = app.sort_record.culture
            culture_id = culture.id
            
            if culture_id not in cultures_data:
                cultures_data[culture_id] = {
                    'culture_id': culture.id,
                    'culture_name': culture.name,
                    'applications': [],
                    'sample_applications': []
                }
            
            cultures_data[culture_id]['applications'].append(app)
            if len(cultures_data[culture_id]['sample_applications']) < 3:
                cultures_data[culture_id]['sample_applications'].append(
                    app.application_number
                )
        
        # Подсчитать статистику
        result = []
        for culture_id, data in cultures_data.items():
            apps = data['applications']
            
            # Проверить сколько уже в испытаниях
            in_trial_count = 0
            for app in apps:
                if TrialParticipant.objects.filter(
                    application=app,
                    trial__region=region,
                    is_deleted=False
                ).exists():
                    in_trial_count += 1
            
            result.append({
                'culture_id': culture_id,
                'culture_name': data['culture_name'],
                'applications_count': len(apps),
                'pending_count': len(apps) - in_trial_count,
                'in_trial_count': in_trial_count,
                'sample_applications': data['sample_applications']
            })
        
        # Сортировать по количеству заявок (больше заявок → выше)
        result.sort(key=lambda x: x['applications_count'], reverse=True)
        
        return Response({
            'region_id': region.id,
            'region_name': region.name,
            'oblast_id': region.oblast.id,
            'oblast_name': region.oblast.name,
            'cultures': result
        })
    
    @action(detail=True, methods=['get'], url_path='regional-status')
    def regional_status(self, request, pk=None):
        """
        Получить статусы испытаний по каждой области для заявки
        
        GET /api/v1/applications/{id}/regional-status/
        
        Response:
        {
          "application_id": 5,
          "application_status": "in_progress",
          "regions": [
            {
              "planned_distribution_id": 1,
              "region_id": 1,
              "region_name": "Алматинский ГСУ",
              "oblast_name": "Алматинская",
              "status": "trial_created",
              "trial_id": 15,
              "trial_status": "active"
            },
            {
              "planned_distribution_id": 2,
              "region_id": 5,
              "region_name": "Акмолинский ГСУ",
              "oblast_name": "Акмолинская",
              "status": "planned",
              "trial_id": null,
              "trial_status": null
            }
          ]
        }
        """
        application = self.get_object()
        
        distributions = PlannedDistribution.objects.filter(
            application=application,
            is_deleted=False
        ).select_related('region__oblast')
        
        regions_data = []
        for dist in distributions:
            # Получаем все Trial для этого PlannedDistribution (многолетние)
            trials = dist.get_trials()
            latest_trial = dist.get_latest_trial()
            
            region_info = {
                'planned_distribution_id': dist.id,
                'region_id': dist.region.id,
                'region_name': dist.region.name,
                'oblast_name': dist.region.oblast.name,
                
                # Статус PlannedDistribution
                'status': dist.status,
                'status_display': dist.get_status_display(),
                
                # Многолетняя информация
                'year_started': dist.year_started,
                'year_completed': dist.year_completed,
                'years_count': dist.get_years_count(),
                
                # Последний Trial (если есть)
                'latest_trial_id': latest_trial.id if latest_trial else None,
                'latest_trial_status': latest_trial.status if latest_trial else None,
                'latest_decision': dist.get_latest_decision(),
                
                # Все Trial (список по годам)
                'trials': []
            }
            
            # Добавляем информацию о всех Trial
            for trial in trials:
                trial_info = {
                    'trial_id': trial.id,
                    'year': trial.year or (trial.start_date.year if trial.start_date else None),
                    'start_date': trial.start_date,
                    'status': trial.status,
                    'status_display': trial.get_status_display(),
                    'decision': trial.decision,
                    'decision_display': trial.get_decision_display() if trial.decision else None,
                    'participants_count': trial.participants.filter(is_deleted=False).count(),
                    'laboratory_status': trial.laboratory_status,
                    'responsible_person': trial.responsible_person,
                }
                region_info['trials'].append(trial_info)
            
            regions_data.append(region_info)
        
        return Response({
            'application_id': application.id,
            'application_number': application.application_number,
            'application_status': application.status,
            'total_regions': len(regions_data),
            'regions': regions_data
        })
    
    @action(detail=False, methods=['get'], url_path='pending-for-region')
    def pending_for_region(self, request):
        """
        Получить распределенные заявки для ГСУ по культуре
        
        Используется сортопытом при добавлении участников.
        
        GET /api/v1/applications/pending-for-region/?region_id=1&culture_id=1
        GET /api/v1/applications/pending-for-region/?region_id=1  (все культуры)
        """
        culture_id = request.query_params.get('culture_id')
        region_id = request.query_params.get('region_id')
        trial_id = request.query_params.get('trial_id')
        
        # region_id обязателен, culture_id - опционально
        if not region_id:
            return Response({
                'error': 'region_id is required'
            }, status=400)
        
        try:
            region = Region.objects.get(id=region_id, is_deleted=False)
        except Region.DoesNotExist:
            return Response({'error': 'Region not found'}, status=404)
        
        # Базовый фильтр - распределенные заявки для этой области
        applications_qs = Application.objects.filter(
            status__in=['distributed', 'in_progress'],
            target_oblasts__id=region.oblast_id,
            is_deleted=False
        )
        
        # Фильтр по культуре (если указана)
        if culture_id:
            applications_qs = applications_qs.filter(
                sort_record__culture_id=culture_id
            )
        
        applications_qs = applications_qs.distinct()
        
        result = []
        for app in applications_qs:
            app_data = {
                'id': app.id,
                'application_number': app.application_number,
                'sort_record': SortRecordSerializer(app.sort_record).data if app.sort_record else None,
                'already_in_trial': False
            }
            
            # Проверить добавлен ли в указанное испытание
            if trial_id:
                app_data['already_in_trial'] = TrialParticipant.objects.filter(
                    trial_id=trial_id,
                    sort_record=app.sort_record,
                    is_deleted=False
                ).exists()
            
            result.append(app_data)
        
        return Response({
            'total': len(result),
            'region_id': region.id,
            'region_name': region.name,
            'culture_id': culture_id if culture_id else None,
            'applications': result
        })


class TrialViewSet(viewsets.ModelViewSet):
    """
    Управление испытаниями сортов
    
    Сорта связываются через sort_id из Patents Service
    Культуры получаются через API (не хранятся локально)
    """
    queryset = Trial.objects.filter(is_deleted=False)
    serializer_class = TrialSerializer
    
    def get_permissions(self):
        """Чтение, список сортов - всем, остальное - только авторизованным"""
        if self.action in ['list', 'retrieve', 'available_sorts', 'form008', 'form008_statistics']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]
    
    def perform_create(self, serializer):
        """При создании испытания устанавливаем created_by"""
        from django.core.exceptions import ObjectDoesNotExist
        from rest_framework.exceptions import ValidationError
        
        try:
            serializer.save(created_by=self.request.user)
        except ObjectDoesNotExist as e:
            # Понятное сообщение об отсутствующем объекте
            raise ValidationError({
                'error': f'Связанный объект не найден: {str(e)}'
            })
        except Exception as e:
            # Логируем ошибку для отладки
            import traceback
            error_msg = str(e)
            print(f"Error creating trial: {error_msg}")
            print(traceback.format_exc())
            
            # Более понятное сообщение для пользователя
            if 'foreign key constraint' in error_msg.lower() or 'does not exist' in error_msg.lower():
                raise ValidationError({
                    'error': 'Один или несколько указанных ID не существуют в базе данных',
                    'details': error_msg
                })
            raise
    
    @action(detail=False, methods=['get'], url_path='available-sorts')
    def available_sorts(self, request):
        """
        Получить список доступных сортов из Patents Service
        
        Проксирует запрос к Patents Service API для получения списка сортов.
        Используется фронтендом для выбора сорта при создании испытания.
        
        GET /api/v1/trials/available-sorts/
        """
        sorts = patents_api.get_all_sorts(params=request.query_params.dict())
        return Response(sorts)
    
    @action(detail=False, methods=['post'], url_path='validate-sort')
    def validate_sort(self, request):
        """
        Валидировать сорт перед созданием испытания
        
        Проверяет существование сорта в Patents Service и возвращает
        данные для денормализации.
        
        POST /api/v1/trials/validate-sort/
        Body: {"sort_id": 456}
        
        Returns:
            {
                "valid": true,
                "data": {
                    "sort_id": 456,
                    "sort_name": "Пшеница Акмола 3",
                    "culture_id": 1,
                    "culture_name": "Пшеница яровая"
                }
            }
        """
        sort_id = request.data.get('sort_id')
        if not sort_id:
            return Response({
                'valid': False,
                'error': 'sort_id required'
            }, status=400)
        
        sort_data = patents_api.validate_sort_for_trial(sort_id)
        if sort_data:
            return Response({
                'valid': True,
                'data': sort_data
            })
        else:
            return Response({
                'valid': False,
                'error': f'Sort with ID {sort_id} not found in Patents Service'
            }, status=404)
    
    @action(detail=True, methods=['post'], url_path='decision')
    def make_decision(self, request, pk=None):
        """
        Принять решение по испытанию
        
        После принятия решения автоматически обновляется статус заявки.
        
        POST /api/v1/trials/{id}/decision/
        Body: {
            "decision": "approved",  # approved | continue | rejected
            "justification": "Сорт показал отличные результаты...",
            "recommendations": "Рекомендовать к включению в реестр",
            "decision_date": "2025-10-01"
        }
        """
        trial = self.get_object()
        
        # Проверяем что испытание завершено
        if trial.status not in ['completed', 'lab_completed', 'completed_008']:
            return Response({
                'error': 'Trial must be completed before making decision'
            }, status=400)
        
        decision = request.data.get('decision')
        if decision not in ['approved', 'continue', 'rejected']:
            return Response({
                'error': 'Invalid decision. Must be: approved, continue, or rejected'
            }, status=400)
        
        # Сохраняем решение
        trial.decision = decision
        trial.decision_justification = request.data.get('justification', '')
        trial.decision_recommendations = request.data.get('recommendations', '')
        trial.decision_date = request.data.get('decision_date', timezone.now().date())
        trial.decided_by = request.user if request.user.is_authenticated else None
        
        # Обновляем статус испытания
        trial.status = decision  # approved, continue, или rejected
        trial.save()
        
        # АВТОМАТИЧЕСКОЕ ОБНОВЛЕНИЕ PlannedDistribution (многолетние испытания):
        # Ищем PlannedDistribution через участников
        from trials_app.models import TrialParticipant
        participants_with_app = TrialParticipant.objects.filter(
            trial=trial,
            application__isnull=False,
            is_deleted=False
        ).select_related('application')
        
        for participant in participants_with_app:
            planned_dist = PlannedDistribution.objects.filter(
                application=participant.application,
                region=trial.region,
                status='in_progress'
            ).first()
            
            if planned_dist:
                # Финальное решение: approved или rejected → завершаем PlannedDistribution
                if decision in ['approved', 'rejected']:
                    planned_dist.status = decision  # approved или rejected
                    planned_dist.year_completed = trial.year or (trial.start_date.year if trial.start_date else None)
                    planned_dist.save()
                # decision = 'continue' → оставляем in_progress (будет новый Trial в следующем году)
        
        return Response({
            'success': True,
            'message': f'Decision "{decision}" saved successfully',
            'trial': TrialSerializer(trial).data
        })
    
    @action(detail=True, methods=['post'], url_path='mark-sent-to-lab')
    def mark_sent_to_lab(self, request, pk=None):
        """
        Отметить что образец отправлен в лабораторию
        
        POST /api/v1/trials/{id}/mark-sent-to-lab/
        Body: {
            "laboratory_code": "LAB-2025-001-ALM",
            "sample_weight_kg": 2.0,
            "sent_date": "2025-10-15",
            "participant_id": 50,  # опционально - от какого участника образец
            "sample_source": "Образец из делянки №2"
        }
        """
        trial = self.get_object()
        
        # Проверка что испытание в статусе completed_008
        if trial.status != 'completed_008':
            return Response({
                'error': f'Trial must be in status completed_008 before sending to lab. Current status: {trial.status}'
            }, status=400)
        
        # Обновляем данные
        trial.laboratory_status = 'sent'
        trial.laboratory_code = request.data.get('laboratory_code')
        trial.laboratory_sent_date = request.data.get('sent_date', timezone.now().date())
        trial.laboratory_sample_weight = request.data.get('sample_weight_kg')
        trial.laboratory_sample_source = request.data.get('sample_source', '')
        trial.laboratory_notes = request.data.get('notes', '')
        trial.status = 'lab_sample_sent'
        trial.save()
        
        # АВТОМАТИЧЕСКИ СОЗДАТЬ КАЧЕСТВЕННЫЕ ПОКАЗАТЕЛИ ДЛЯ ЛАБОРАТОРИИ
        quality_results = create_quality_trial_results(trial, request.user)
        
        return Response({
            'success': True,
            'message': f'Sample {trial.laboratory_code} marked as sent to laboratory',
            'quality_indicators_created': len(quality_results),
            'trial': TrialSerializer(trial).data
        })
    
    @action(detail=True, methods=['post'], url_path='laboratory-results/bulk-entry')
    def laboratory_results_bulk_entry(self, request, pk=None):
        """
        Массовое внесение лабораторных результатов
        
        POST /api/v1/trials/{id}/laboratory-results/bulk-entry/
        Body: {
            "laboratory_code": "LAB-2025-001-ALM",
            "analysis_date": "2025-10-20",
            "participant_id": 50,  # опционально - для какого участника
            "results": [
                {
                    "indicator": 10,  # ID показателя (должен быть is_quality=True)
                    "value": 14.5
                },
                {
                    "indicator": 11,
                    "value": 28.0
                },
                {
                    "indicator": 12,
                    "value": 785
                }
            ]
        }
        """
        trial = self.get_object()
        
        # Проверка статуса
        if trial.laboratory_status != 'sent':
            return Response({
                'error': f'Trial must have laboratory_status=sent. Current: {trial.laboratory_status}'
            }, status=400)
        
        laboratory_code = request.data.get('laboratory_code')
        analysis_date = request.data.get('analysis_date', timezone.now().date())
        participant_id = request.data.get('participant_id')
        results_data = request.data.get('results', [])
        
        if not results_data:
            return Response({
                'error': 'No results provided'
            }, status=400)
        
        # Проверить participant если указан
        participant = None
        if participant_id:
            try:
                participant = TrialParticipant.objects.get(id=participant_id, trial=trial)
            except TrialParticipant.DoesNotExist:
                return Response({
                    'error': f'Participant {participant_id} not found in this trial'
                }, status=404)
        
        created_results = []
        updated_results = []
        
        for item in results_data:
            indicator_id = item.get('indicator')
            value = item.get('value')
            text_value = item.get('text_value')
            
            if not indicator_id:
                continue
            
            # Проверить что показатель существует и is_quality=True
            try:
                indicator = Indicator.objects.get(id=indicator_id)
                if not indicator.is_quality:
                    return Response({
                        'error': f'Indicator "{indicator.name}" is not a quality indicator (is_quality=False)'
                    }, status=400)
            except Indicator.DoesNotExist:
                return Response({
                    'error': f'Indicator {indicator_id} not found'
                }, status=404)
            
            # Создать или обновить результат
            result, created = TrialLaboratoryResult.objects.update_or_create(
                trial=trial,
                indicator=indicator,
                participant=participant,
                defaults={
                    'value': value,
                    'text_value': text_value,
                    'laboratory_code': laboratory_code,
                    'analysis_date': analysis_date,
                    'created_by': request.user
                }
            )
            
            if created:
                created_results.append(result)
            else:
                updated_results.append(result)
        
        return Response({
            'success': True,
            'created': len(created_results),
            'updated': len(updated_results),
            'total': len(results_data)
        })
    
    @action(detail=True, methods=['post'], url_path='complete')
    def complete(self, request, pk=None):
        """
        Завершить полевые работы
        
        POST /api/trials/{id}/complete/
        Body: {
            "completed_date": "2025-10-22"  # опционально
        }
        
        Переводит испытание в статус 'completed' (испытание полностью завершено).
        Это финальный статус после всех этапов.
        """
        trial = self.get_object()
        
        # Проверка текущего статуса - можно завершить из любого статуса
        if trial.status in ['approved', 'continue', 'rejected']:
            return Response({
                'error': f'Trial already has final decision. Current: {trial.status}'
            }, status=400)
        
        # Обновляем статус
        trial.status = 'completed'
        if 'completed_date' in request.data:
            # Если пользователь хочет установить кастомную дату
            from datetime.datetime import strptime
            completed_date = request.data['completed_date']
        else:
            completed_date = None
        
        trial.save()
        
        return Response({
            'success': True,
            'message': 'Trial completed successfully',
            'trial': TrialSerializer(trial).data
        })
    
    @action(detail=True, methods=['post'], url_path='laboratory-complete')
    def laboratory_complete(self, request, pk=None):
        """
        Завершить лабораторные анализы
        
        POST /api/v1/trials/{id}/laboratory-complete/
        Body: {
            "completed_date": "2025-10-22"  # опционально
        }
        """
        trial = self.get_object()
        
        # Проверка статуса
        if trial.laboratory_status != 'sent':
            return Response({
                'error': f'Trial must have laboratory_status=sent. Current: {trial.laboratory_status}'
            }, status=400)
        
        # Обновляем статусы
        trial.laboratory_status = 'completed'
        trial.laboratory_completed_date = request.data.get('completed_date', timezone.now().date())
        trial.status = 'lab_completed'
        trial.save()
        
        return Response({
            'success': True,
            'message': 'Laboratory analyses completed',
            'trial': TrialSerializer(trial).data
        })
    
    @action(detail=True, methods=['post'], url_path='lcomplete')
    def lcomplete(self, request, pk=None):
        """
        Завершить испытание (финальное завершение после лабораторных анализов)
        
        POST /api/trials/{id}/lcomplete/
        Body: {
            "completed_date": "2025-10-22"  # опционально
        }
        
        Переводит испытание из статуса 'lab_completed' в финальный статус 'completed'.
        После этого можно принимать решение комиссии.
        """
        trial = self.get_object()
        
        # Проверка текущего статуса
        if trial.status != 'lab_completed':
            return Response({
                'error': f'Trial must be in status lab_completed. Current: {trial.status}'
            }, status=400)
        
        # Обновляем статус на финальный completed
        trial.status = 'completed'
        trial.save()
        
        return Response({
            'success': True,
            'message': 'Trial completed successfully',
            'trial': TrialSerializer(trial).data
        })
    
    @action(detail=True, methods=['get'], url_path='laboratory-results')
    def get_laboratory_results(self, request, pk=None):
        """
        Получить лабораторные результаты испытания
        
        GET /api/v1/trials/{id}/laboratory-results/
        """
        trial = self.get_object()
        
        results = TrialLaboratoryResult.objects.filter(
            trial=trial,
            is_deleted=False
        )
        
        serializer = TrialLaboratoryResultSerializer(results, many=True)
        
        return Response({
            'trial_id': trial.id,
            'laboratory_status': trial.laboratory_status,
            'laboratory_code': trial.laboratory_code,
            'results_count': results.count(),
            'results': serializer.data
        })
    
    @action(detail=True, methods=['get'], url_path='form008')
    def form008(self, request, pk=None):
        """
        Получить форму 008 для заполнения результатов испытания
        
        GET /api/v1/trials/{id}/form008/
        
        Возвращает структуру для массового внесения результатов
        по всем участникам и показателям.
        """
        trial = self.get_object()
        
        # Получить участников
        participants = trial.participants.filter(is_deleted=False).order_by('participant_number')
        
        participants_data = []
        for participant in participants:
            # Получить текущие результаты участника
            results = TrialResult.objects.filter(
                participant=participant,
                is_deleted=False
            ).select_related('indicator')
            
            current_results = {}
            for result in results:
                current_results[result.indicator.code] = {
                    'value': result.value,
                    'text_value': result.text_value,
                    'measurement_date': result.measurement_date
                }
            
            participants_data.append({
                'id': participant.id,
                'participant_number': participant.participant_number,
                'sort_name': participant.sort_record.name if participant.sort_record else None,
                'sort_code': participant.sort_record.public_code if participant.sort_record else None,
                'statistical_group': participant.statistical_group,
                'statistical_result': participant.statistical_result,
                'is_standard': participant.is_standard,
                'application_number': participant.application.application_number if participant.application else None,
                'current_results': current_results
            })
        
        # Получить показатели для культуры
        indicators_data = []
        if trial.culture:
            indicators = trial.indicators.filter(
                is_deleted=False,
                is_quality=False  # Только основные показатели (не лабораторные)
            ).order_by('sort_order', 'name')
            
            for indicator in indicators:
                indicators_data.append({
                    'id': indicator.id,
                    'code': indicator.code,
                    'name': indicator.name,
                    'unit': indicator.unit,
                    'is_numeric': indicator.is_numeric,
                    'sort_order': indicator.sort_order
                })
        
        # Рассчитать min/max по каждому показателю
        min_max = {}
        for indicator in indicators:
            values = TrialResult.objects.filter(
                participant__trial=trial,
                indicator=indicator,
                value__isnull=False,
                is_deleted=False
            ).values_list('value', flat=True)
            
            if values:
                min_max[indicator.code] = {
                    'min': min(values),
                    'max': max(values)
                }
            else:
                min_max[indicator.code] = {
                    'min': None,
                    'max': None
                }
        
        # Получить статистику
        statistics = trial.calculate_trial_statistics()
        
        return Response({
            'trial': {
                'id': trial.id,
                'year': trial.year or (trial.start_date.year if trial.start_date else None),
                'region_name': trial.region.name,
                'region_code': None,  # TODO: добавить код региона в модель
                'oblast_name': trial.region.oblast.name,
                'culture_name': trial.culture.name if trial.culture else None,
                'culture_group': trial.culture.group_culture.name if trial.culture and trial.culture.group_culture else None,
                'trial_type': trial.trial_type.name if trial.trial_type else None,
                'trial_type_code': trial.trial_type.code if trial.trial_type else None,
                'predecessor': self._get_predecessor_display(trial),
                'growing_conditions': trial.get_growing_conditions_display() if trial.growing_conditions else None,
                'cultivation_technology': trial.get_cultivation_technology_display() if trial.cultivation_technology else None,
                'growing_method': trial.get_growing_method_display() if trial.growing_method else None,
                'harvest_timing': trial.get_harvest_timing_display() if trial.harvest_timing else None,
                'harvest_date': trial.harvest_date,
                'status': trial.status,
                'status_display': trial.get_status_display()
            },
            'participants': participants_data,
            'indicators': indicators_data,
            'statistics': statistics or {
                'calculated': False,
                'sx': None,
                'accuracy_percent': None,
                'lsd': None,
                'error_mean': None
            },
            'min_max': min_max
        })
    
    def _get_predecessor_display(self, trial):
        """Получить отображение предшественника"""
        if trial.predecessor_culture:
            return trial.predecessor_culture.name
        return "пар"
    
    @action(detail=True, methods=['post'], url_path='form008/bulk-save')
    def form008_bulk_save(self, request, pk=None):
        """
        Массовое сохранение результатов формы 008
        
        POST /api/v1/trials/{id}/form008/bulk-save/
        {
            "is_final": false,  // false = черновик, true = финальная отправка
            "harvest_date": "2024-09-23",
            "measurement_date": "2024-09-23",
            "participants": [
                {
                    "participant_id": 50,
                    "results": {
                        "yield": 7.9,
                        "seed_weight_1000": 32.9,
                        "lodging_resistance": 4,
                        ...
                    }
                }
            ]
        }
        """
        trial = self.get_object()
        
        # Проверка статуса
        if trial.status not in ['active', 'planned']:
            return Response({
                'error': f'Cannot save results. Trial status is {trial.status}'
            }, status=400)
        
        is_final = request.data.get('is_final', False)
        harvest_date = request.data.get('harvest_date')
        measurement_date = request.data.get('measurement_date', harvest_date)
        participants_data = request.data.get('participants', [])
        
        if not participants_data:
            return Response({
                'error': 'No participants data provided'
            }, status=400)
        
        # Обновить harvest_date в trial
        if harvest_date:
            trial.harvest_date = harvest_date
            trial.save()
        
        results_created = 0
        results_updated = 0
        
        # Сохранить результаты для каждого участника
        for p_data in participants_data:
            participant_id = p_data.get('participant_id')
            results = p_data.get('results', {})
            
            if not participant_id:
                continue
            
            try:
                participant = TrialParticipant.objects.get(id=participant_id, trial=trial)
            except TrialParticipant.DoesNotExist:
                continue
            
            # Сохранить каждый показатель
            for indicator_code, value in results.items():
                # Найти показатель по коду
                try:
                    indicator = Indicator.objects.get(code=indicator_code, is_deleted=False)
                except Indicator.DoesNotExist:
                    continue
                
                # Создать или обновить результат
                result_obj, created = TrialResult.objects.update_or_create(
                    participant=participant,
                    indicator=indicator,
                    defaults={
                        'value': value if isinstance(value, (int, float)) else None,
                        'text_value': str(value) if not isinstance(value, (int, float)) else None,
                        'measurement_date': measurement_date,
                        'trial': trial,
                        'sort_record': participant.sort_record,
                        'created_by': request.user
                    }
                )
                
                if created:
                    results_created += 1
                else:
                    results_updated += 1
        
        # Пересчитать статистику и statistical_result для каждого участника
        statistics = trial.calculate_trial_statistics()
        
        # Обновить statistical_result для всех участников
        for participant in trial.participants.filter(is_deleted=False):
            participant.calculate_statistical_result()
        
        # Рассчитать min/max
        min_max = {}
        for indicator in trial.indicators.filter(is_deleted=False, is_quality=False):
            values = TrialResult.objects.filter(
                participant__trial=trial,
                indicator=indicator,
                value__isnull=False,
                is_deleted=False
            ).values_list('value', flat=True)
            
            if values:
                min_max[indicator.code] = {
                    'min': float(min(values)),
                    'max': float(max(values))
                }
        
        # Если финальная отправка - изменить статус
        if is_final:
            trial.status = 'completed_008'
            trial.save()
            message = 'Form 008 submitted successfully. Field work completed.'
        else:
            message = 'Form 008 draft saved successfully'
        
        # Получить стандарт для расчета отклонений
        standard_participant = trial.get_standard_participants().first()
        standard_yield = None
        
        if standard_participant:
            standard_yield_result = TrialResult.objects.filter(
                participant=standard_participant,
                indicator__code='yield',
                is_deleted=False
            ).first()
            if standard_yield_result:
                standard_yield = standard_yield_result.value
        
        # Получить statistical_result для всех участников с отклонениями
        participants_stats = []
        for participant in trial.participants.filter(is_deleted=False).order_by('participant_number'):
            # Получить урожайность
            yield_result = TrialResult.objects.filter(
                participant=participant,
                indicator__code='yield',
                is_deleted=False
            ).first()
            
            participant_yield = yield_result.value if yield_result else None
            
            # Рассчитать отклонения от стандарта
            deviation_abs = None
            deviation_pct = None
            
            if participant_yield is not None and standard_yield is not None and not participant.is_standard:
                deviation_abs = round(participant_yield - standard_yield, 2)
                deviation_pct = round((deviation_abs / standard_yield * 100), 1) if standard_yield != 0 else None
            
            participants_stats.append({
                'participant_id': participant.id,
                'participant_number': participant.participant_number,
                'sort_name': participant.sort_record.name if participant.sort_record else None,
                'yield': participant_yield,
                'statistical_result': participant.statistical_result,
                'statistical_result_display': participant.get_statistical_result_display() if participant.statistical_result is not None else None,
                'deviation_standard_abs': deviation_abs,
                'deviation_standard_pct': deviation_pct,
                'is_standard': participant.is_standard
            })
        
        return Response({
            'success': True,
            'message': message,
            'is_final': is_final,
            'results_created': results_created,
            'results_updated': results_updated,
            'trial_status': trial.status,
            'statistics': statistics or {},
            'min_max': min_max,
            'participants_statistical_results': participants_stats
        })
    
    @action(detail=True, methods=['get'], url_path='form008/statistics')
    def form008_statistics(self, request, pk=None):
        """
        Получить статистику испытания (P%, НСР, E)
        
        GET /api/v1/trials/{id}/form008/statistics/
        """
        trial = self.get_object()
        
        statistics = trial.calculate_trial_statistics()
        
        if not statistics:
            return Response({
                'has_data': False,
                'message': 'No data available for statistics calculation'
            })
        
        # Найти стандарт
        standard = trial.get_standard_participants().first()
        standard_data = None
        
        if standard:
            yield_result = TrialResult.objects.filter(
                participant=standard,
                indicator__code='yield',
                is_deleted=False
            ).first()
            
            if yield_result:
                standard_data = {
                    'participant_id': standard.id,
                    'sort_name': standard.sort_record.name if standard.sort_record else None,
                    'yield': yield_result.value,
                    'plot_values': [
                        yield_result.plot_1,
                        yield_result.plot_2,
                        yield_result.plot_3,
                        yield_result.plot_4
                    ] if yield_result.plot_1 is not None else None
                }
        
        # Сравнение всех участников со стандартом
        comparison = []
        for participant in trial.get_tested_participants():
            yield_result = TrialResult.objects.filter(
                participant=participant,
                indicator__code='yield',
                is_deleted=False
            ).first()
            
            if yield_result and standard_data:
                deviation_abs = yield_result.value - standard_data['yield']
                deviation_pct = (deviation_abs / standard_data['yield'] * 100) if standard_data['yield'] != 0 else None
                
                comparison.append({
                    'participant_id': participant.id,
                    'participant_number': participant.participant_number,
                    'sort_name': participant.sort_record.name if participant.sort_record else None,
                    'yield': yield_result.value,
                    'deviation_standard_abs': round(deviation_abs, 2),
                    'deviation_standard_pct': round(deviation_pct, 1) if deviation_pct is not None else None,
                    'statistical_result': participant.statistical_result,
                    'statistical_result_display': participant.get_statistical_result_display() if participant.statistical_result is not None else None
                })
        
        return Response({
            'trial_id': trial.id,
            'has_data': True,
            'statistics': statistics,
            'standard': standard_data,
            'comparison': comparison
        })

class TrialParticipantViewSet(viewsets.ModelViewSet):
    """Участники сортоопытов"""
    queryset = TrialParticipant.objects.filter(is_deleted=False)
    serializer_class = TrialParticipantSerializer
    permission_classes = [permissions.IsAuthenticated]  # Требуется авторизация
    
    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        """
        Массовое создание участников для испытания
        
        POST /api/v1/trial-participants/bulk-create/
        {
            "trial": 10,
            "participants": [
                {
                    "sort_record": 100,
                    "statistical_group": 0,
                    "participant_number": 1
                },
                {
                    "sort_record": 50,
                    "statistical_group": 1,
                    "participant_number": 2,
                    "application": 5
                }
            ]
        }
        """
        trial_id = request.data.get('trial')
        participants_data = request.data.get('participants', [])
        
        if not trial_id:
            return Response({'error': 'trial is required'}, status=400)
        
        try:
            trial = Trial.objects.get(id=trial_id)
        except Trial.DoesNotExist:
            return Response({'error': 'Trial not found'}, status=404)
        
        created_participants = []
        for p_data in participants_data:
            participant = TrialParticipant.objects.create(
                trial=trial,
                sort_record_id=p_data['sort_record'],
                statistical_group=p_data.get('statistical_group', 1),
                participant_number=p_data['participant_number'],
                application_id=p_data.get('application'),
            )
            created_participants.append(participant)
            
            # АВТОМАТИЧЕСКИ СОЗДАТЬ ОСНОВНЫЕ ПОКАЗАТЕЛИ ДЛЯ СОРТОПЫТА
            create_basic_trial_results(participant, request.user)
        
        # АВТОМАТИЧЕСКИЙ СТАТУС TRIAL:
        # Если Trial был planned и добавили участников → переводим в active
        if trial.status == 'planned' and created_participants:
            trial.status = 'active'
            trial.save()
        
        # АВТОМАТИЧЕСКОЕ ОБНОВЛЕНИЕ СТАТУСА ЗАЯВКИ И PlannedDistribution:
        application_ids = set()
        for participant in created_participants:
            if participant.application_id:
                application_ids.add(participant.application_id)
        
        if application_ids:
            # Обновляем общий статус заявки
            Application.objects.filter(
                id__in=application_ids,
                status='distributed'
            ).update(status='in_progress')
            
            # Обновляем PlannedDistribution (многолетние испытания)
            for app_id in application_ids:
                planned_dist = PlannedDistribution.objects.filter(
                    application_id=app_id,
                    region=trial.region
                ).first()
                
                if planned_dist:
                    # Если первый Trial для этого PlannedDistribution
                    if planned_dist.status == 'planned':
                        planned_dist.status = 'in_progress'
                        planned_dist.year_started = trial.year or (trial.start_date.year if trial.start_date else None)
                        planned_dist.save()
        
        serializer = TrialParticipantSerializer(created_participants, many=True)
        return Response({
            'success': True,
            'count': len(created_participants),
            'participants': serializer.data,
            'trial_status': trial.status,
            'updated_applications': list(application_ids)
        })


class TrialResultViewSet(viewsets.ModelViewSet):
    """Результаты испытаний"""
    queryset = TrialResult.objects.filter(is_deleted=False)
    serializer_class = TrialResultSerializer
    permission_classes = [permissions.IsAuthenticated]  # Требуется авторизация
    
    @action(detail=False, methods=['post'], url_path='bulk-entry')
    def bulk_entry(self, request):
        """
        Массовое внесение результатов (для сортопыта)
        
        POST /api/v1/trial-results/bulk-entry/
        {
            "trial": 10,
            "participant": 5,
            "measurement_date": "2024-09-23",
            "data": [
                {
                    "indicator": 1,  // урожайность
                    "plots": [45.2, 46.8, 44.5, 45.9]
                },
                {
                    "indicator": 2,  // белок
                    "value": 14.5
                },
                {
                    "indicator": 3,  // устойчивость к полеганию
                    "value": 5
                }
            ]
        }
        """
        participant_id = request.data.get('participant')
        measurement_date = request.data.get('measurement_date')
        data = request.data.get('data', [])
        
        if not participant_id:
            return Response({'error': 'participant is required'}, status=400)
        
        try:
            participant = TrialParticipant.objects.get(id=participant_id)
        except TrialParticipant.DoesNotExist:
            return Response({'error': 'Participant not found'}, status=404)
        
        created_results = []
        updated_results = []
        
        for item in data:
            indicator_id = item['indicator']
            
            # Проверить, существует ли результат (обычно уже создан автоматически)
            result, created = TrialResult.objects.get_or_create(
                participant=participant,
                indicator_id=indicator_id,
                defaults={
                    'trial': participant.trial,
                    'sort_record': participant.sort_record,
                    'measurement_date': measurement_date,
                    'created_by': request.user
                }
            )
            
            # Обновить данные (работает как для существующих, так и для новых записей)
            if 'plots' in item and len(item['plots']) == 4:
                result.plot_1 = item['plots'][0]
                result.plot_2 = item['plots'][1]
                result.plot_3 = item['plots'][2]
                result.plot_4 = item['plots'][3]
                # Можно сбросить value если заполняются делянки
                # (среднее рассчитается автоматически в модели)
            elif 'value' in item:
                result.value = item['value']
                # Если передается итоговое значение, очищаем делянки
                result.plot_1 = None
                result.plot_2 = None
                result.plot_3 = None
                result.plot_4 = None
            
            if 'text_value' in item:
                result.text_value = item['text_value']
            
            if measurement_date:
                result.measurement_date = measurement_date
            
            result.save()
            
            if created:
                created_results.append(result)
            else:
                updated_results.append(result)
        
        return Response({
            'success': True,
            'created': len(created_results),
            'updated': len(updated_results),
            'total': len(data)
        })

class DocumentViewSet(viewsets.ModelViewSet):
    """Документы испытаний и заявок"""
    queryset = Document.objects.filter(is_deleted=False)
    serializer_class = DocumentSerializer
    permission_classes = [permissions.IsAuthenticated]  # Требуется авторизация
    
    def get_queryset(self):
        """Фильтрация по application или trial"""
        queryset = super().get_queryset()
        
        application_id = self.request.query_params.get('application')
        trial_id = self.request.query_params.get('trial')
        
        if application_id:
            queryset = queryset.filter(application_id=application_id)
        
        if trial_id:
            queryset = queryset.filter(trial_id=trial_id)
        
        return queryset
    
    def perform_create(self, serializer):
        """При создании документа устанавливаем uploaded_by автоматически"""
        serializer.save(uploaded_by=self.request.user)
    
    @action(detail=True, methods=['get'], url_path='download')
    def download(self, request, pk=None):
        """
        Скачать файл документа
        
        GET /api/v1/documents/{id}/download/
        
        Возвращает файл для скачивания
        """
        from django.http import FileResponse, Http404
        import os
        
        document = self.get_object()
        
        if not document.file:
            return Response({
                'error': 'Document has no file attached'
            }, status=404)
        
        try:
            file_path = document.file.path
            if not os.path.exists(file_path):
                return Response({
                    'error': 'File not found on server'
                }, status=404)
            
            # Открываем файл и возвращаем как attachment
            response = FileResponse(
                open(file_path, 'rb'),
                as_attachment=True,
                filename=os.path.basename(file_path)
            )
            return response
        except Exception as e:
            return Response({
                'error': f'Error downloading file: {str(e)}'
            }, status=500)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_cultures(request):
    """
    Получить список всех культур из Patents Service
    
    Проксирует запрос к Patents Service для получения справочника культур.
    Используется при создании/редактировании сортов.
    
    GET /api/v1/patents/cultures/
    
    Query params:
        - group: фильтр по группе культур (ID группы)
        - group_culture_id: алиас для параметра group (для обратной совместимости)
        - culture_group: алиас для параметра group (для обратной совместимости)
        - search: поиск по названию культуры
    
    Примеры:
        GET /api/v1/patents/cultures/?group=1  # Зерновые культуры
        GET /api/v1/patents/cultures/?search=пшеница  # Поиск по названию
        GET /api/v1/patents/cultures/?group=1&search=пшеница  # Комбинированная фильтрация
    """
    cultures = patents_api.get_all_cultures(params=request.query_params.dict())
    return Response(cultures)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_group_cultures(request):
    """
    Получить список всех групп культур из Patents Service
    
    Проксирует запрос к Patents Service для получения справочника групп культур.
    Используется при создании/редактировании культур и сортов.
    
    GET /api/v1/patents/group-cultures/
    
    Query params:
        - search: поиск по названию
    """
    group_cultures = patents_api.get_all_group_cultures(params=request.query_params.dict())
    return Response(group_cultures)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_culture_detail(request, culture_id):
    """
    Получить детальную информацию о культуре
    
    GET /api/v1/patents/cultures/{id}/
    """
    culture = patents_api.get_culture(culture_id)
    if culture:
        return Response(culture)
    else:
        return Response({
            'error': f'Culture {culture_id} not found'
        }, status=404)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def create_culture(request):
    """
    Создать новую культуру в Patents Service
    
    Проксирует запрос к Patents Service.
    Trials Service НЕ хранит культуры локально!
    
    POST /api/v1/patents/cultures/create/
    Body: {
        "name": "Новая культура",
        "group_culture": 1,
        "description": "Описание"
    }
    """
    culture = patents_api.create_culture(request.data)
    if culture:
        return Response(culture, status=201)
    else:
        return Response({
            'error': 'Failed to create culture in Patents Service'
        }, status=500)


@api_view(['PUT', 'PATCH'])
@permission_classes([permissions.AllowAny])
def update_culture(request, culture_id):
    """
    Обновить культуру в Patents Service
    
    PUT/PATCH /api/v1/patents/cultures/{id}/update/
    """
    culture = patents_api.update_culture(culture_id, request.data)
    if culture:
        return Response(culture)
    else:
        return Response({
            'error': f'Failed to update culture {culture_id}'
        }, status=500)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def create_group_culture(request):
    """
    Создать новую группу культур в Patents Service
    
    POST /api/v1/patents/group-cultures/create/
    Body: {
        "name": "Новая группа",
        "description": "Описание"
    }
    """
    group = patents_api.create_group_culture(request.data)
    if group:
        return Response(group, status=201)
    else:
        return Response({
            'error': 'Failed to create group culture in Patents Service'
        }, status=500)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_originators(request):
    """
    Получить список всех оригинаторов из Patents Service
    
    Проксирует запрос к Patents Service.
    Используется при создании/редактировании сортов.
    
    GET /api/v1/patents/originators/
    
    Query params:
        - search: поиск по названию
    """
    originators = patents_api.get_all_originators(params=request.query_params.dict())
    return Response(originators)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_originator_detail(request, originator_id):
    """
    Получить информацию об оригинаторе
    
    GET /api/v1/patents/originators/{id}/
    """
    originator = patents_api.get_originator(originator_id)
    if originator:
        return Response(originator)
    else:
        return Response({
            'error': f'Originator {originator_id} not found'
        }, status=404)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def create_originator(request):
    """
    Создать нового оригинатора в Patents Service
    
    POST /api/v1/patents/originators/create/
    Body: {
        "name": "Название оригинатора"
    }
    """
    originator = patents_api.create_originator(request.data)
    if originator:
        return Response(originator, status=201)
    else:
        return Response({
            'error': 'Failed to create originator in Patents Service'
        }, status=500)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_sorts(request):
    """
    Получить список всех сортов из Patents Service
    
    Проксирует запрос к Patents Service для получения справочника сортов.
    
    GET /api/v1/patents/sorts/
    
    Query params:
        - culture: фильтр по культуре (Patents culture ID)
        - search: поиск по названию
        - code: поиск по коду
    """
    sorts = patents_api.get_all_sorts(params=request.query_params.dict())
    return Response(sorts)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_sorts_for_trial_culture(request):
    """
    Получить сорта для культуры из Trials Service
    
    Принимает локальный Trials culture ID и делает правильный маппинг к Patents API.
    
    GET /api/v1/trials/sorts-by-culture/
    
    Query params:
        - culture_id: локальный ID культуры в Trials (обязательно)
        - region_id: ID региона (опционально) - если передан, возвращается первые 4 сорта
        - search: поиск по названию
        - code: поиск по коду
    
    Response:
        {
            "trials_culture_id": 6,           # Локальный ID в Trials
            "trials_culture_name": "Айва",
            "patents_culture_id": 535,        # ID в Patents Service
            "sorts": [                        # Список сортов из Patents (первые 4 если передан region_id)
                {
                    "id": 2096,
                    "name": "Аксинья",
                    "maturity_group": "D01"   # Только если передан region_id
                },
                {
                    "id": 2097,
                    "name": "Другой сорт",
                    "maturity_group": "D01"   # Только если передан region_id
                }
            ],
            "count": 4 или 10                 # Количество сортов (4 если передан region_id)
        }
    """
    trials_culture_id = request.query_params.get('culture_id')
    
    if not trials_culture_id:
        return Response({
            'error': 'culture_id parameter is required'
        }, status=400)
    
    # Найти культуру в Trials Service
    try:
        culture = Culture.objects.get(id=trials_culture_id, is_deleted=False)
    except Culture.DoesNotExist:
        return Response({
            'error': f'Culture with ID {trials_culture_id} not found in Trials Service'
        }, status=404)
    
    # Получить Patents culture ID
    patents_culture_id = culture.culture_id
    
    if not patents_culture_id:
        return Response({
            'error': f'Culture "{culture.name}" does not have Patents culture_id mapping'
        }, status=400)
    
    # Подготовить параметры для Patents API
    patents_params = {
        'culture': patents_culture_id,  # Используем правильный Patents ID!
    }
    
    # Добавить дополнительные параметры фильтрации
    if 'search' in request.query_params:
        patents_params['search'] = request.query_params['search']
    if 'code' in request.query_params:
        patents_params['code'] = request.query_params['code']
    
    # Запросить сорта из Patents API
    sorts = patents_api.get_all_sorts(params=patents_params)
    
    if sorts is None:
        return Response({
            'error': 'Failed to fetch sorts from Patents Service'
        }, status=500)
    
    # Если передан region_id, возвращаем только первый сорт и добавляем maturity_group к каждому сорту
    region_id = request.query_params.get('region_id')
    response_data = {
        'trials_culture_id': culture.id,
        'trials_culture_name': culture.name,
        'patents_culture_id': patents_culture_id,
        'sorts': sorts,
        'count': len(sorts) if isinstance(sorts, list) else 0
    }
    
    if region_id and isinstance(sorts, list) and len(sorts) > 0:
        # Берем первые 4 сорта и добавляем к каждому maturity_group
        first_4_sorts = sorts[:4]  # Берем первые 4 сорта
        for sort in first_4_sorts:
            if isinstance(sort, dict):
                sort['maturity_group'] = "D01"
        response_data['sorts'] = first_4_sorts
        response_data['count'] = len(first_4_sorts)
    
    return Response(response_data)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_sort_detail(request, sort_id):
    """
    Получить детальную информацию о сорте
    
    GET /api/v1/patents/sorts/{id}/
    """
    sort = patents_api.get_sort(sort_id)
    if sort:
        return Response(sort)
    else:
        return Response({
            'error': f'Sort {sort_id} not found'
        }, status=404)


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def create_sort(request):
    """
    Создать новый сорт в Patents Service
    
    POST /api/v1/patents/sorts/create/
    Body: {
        "name": "Название сорта",
        "code": "371/06",
        "culture": 1,
        "originators": [1, 2]
    }
    """
    sort = patents_api.create_sort(request.data)
    if sort:
        return Response(sort, status=201)
    else:
        return Response({
            'error': 'Failed to create sort in Patents Service'
        }, status=500)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def test_patents_connection(request):
    """
    Тестовый endpoint для проверки подключения к Patents Service
    
    Проверяет:
    - Доступность Patents Service
    - Аутентификацию через токен
    - Получение данных
    
    GET /api/v1/patents/test-connection/
    
    Returns:
        {
            "status": "success",
            "patents_service_url": "http://localhost:8000",
            "authentication": "configured",
            "token_obtained": true,
            "test_request": "success",
            "message": "Successfully connected to Patents Service"
        }
    """
    from .patents_integration import patents_api
    
    result = {
        "status": "success",
        "patents_service_url": patents_api.base_url,
        "authentication": "not configured",
        "token_obtained": False,
        "test_request": "not tested",
        "message": ""
    }
    
    # Проверяем настроены ли credentials
    if patents_api.service_username and patents_api.service_password:
        result["authentication"] = "configured"
        
        # Пробуем получить токен
        token = patents_api.get_auth_token()
        if token:
            result["token_obtained"] = True
            result["token_preview"] = f"{token[:10]}...{token[-10:]}" if len(token) > 20 else token
        else:
            result["status"] = "warning"
            result["message"] = "Failed to obtain authentication token"
    else:
        result["status"] = "warning"
        result["message"] = "Patents Service credentials not configured"
    
    # Пробуем сделать тестовый запрос
    try:
        sorts = patents_api.get_all_sorts()
        if sorts is not None:
            result["test_request"] = "success"
            result["sorts_count"] = len(sorts) if isinstance(sorts, list) else 0
            if result["status"] == "success":
                result["message"] = f"Successfully connected to Patents Service. Found {result['sorts_count']} sorts."
        else:
            result["test_request"] = "failed"
            result["status"] = "error"
            result["message"] = "Failed to fetch data from Patents Service"
    except Exception as e:
        result["test_request"] = "error"
        result["status"] = "error"
        result["error"] = str(e)
        result["message"] = f"Exception while testing connection: {str(e)}"
    
    status_code = 200 if result["status"] == "success" else (500 if result["status"] == "error" else 200)
    return Response(result, status=status_code)


class TrialPlanViewSet(viewsets.ModelViewSet):
    """
    Управление планами испытаний
    
    Планы сортоиспытаний на год с участниками (заявки + сорта из реестра).
    Автоматическое предложение заявок по областям и культурам.
    """
    queryset = TrialPlan.objects.filter(is_deleted=False)
    permission_classes = [permissions.IsAuthenticated]  # Требуется авторизация
    
    def get_serializer_class(self):
        """Используем разные сериализаторы для чтения и записи"""
        if self.action in ['create', 'update', 'partial_update']:
            return TrialPlanWriteSerializer
        return TrialPlanSerializer
    
    def perform_create(self, serializer):
        """При создании плана устанавливаем created_by"""
        serializer.save(created_by=self.request.user)
    
    @action(detail=False, methods=['get'], url_path='suggest-applications')
    def suggest_applications(self, request):
        """
        Предложить заявки для плана по областям и культурам
        
        GET /api/v1/trial-plans/suggest-applications/?oblast_id=1&culture_id=5&trial_type_id=1&season=spring
        
        Возвращает заявки со статусом 'submitted' для указанной области и культуры.
        """
        oblast_id = request.query_params.get('oblast_id')
        culture_id = request.query_params.get('culture_id')
        trial_type_id = request.query_params.get('trial_type_id')
        season = request.query_params.get('season', 'spring')  # По умолчанию весна
        
        if not oblast_id or not culture_id:
            return Response({
                'error': 'oblast_id and culture_id are required'
            }, status=400)
        
        try:
            oblast = Oblast.objects.get(id=oblast_id, is_deleted=False)
            culture = Culture.objects.get(id=culture_id, is_deleted=False)
        except (Oblast.DoesNotExist, Culture.DoesNotExist):
            return Response({'error': 'Oblast or Culture not found'}, status=404)
        
        # Найти заявки для этой области и культуры
        applications = Application.objects.filter(
            target_oblasts=oblast,
            sort_record__culture=culture,
            status='submitted',
            is_deleted=False
        ).select_related('sort_record__culture').distinct()
        
        # Дополнительная фильтрация по типу испытания если указан
        if trial_type_id:
            # Здесь можно добавить логику фильтрации по типу испытания
            pass
        
        result = []
        for app in applications:
            app_data = {
                'id': app.id,
                'application_number': app.application_number,
                'sort_record': {
                    'id': app.sort_record.id,
                    'name': app.sort_record.name,
                    'patents_sort_id': app.sort_record.sort_id
                },
                'submission_date': app.submission_date,
                'applicant': app.applicant,
                'maturity_group': app.maturity_group,
            }
            result.append(app_data)
        
        return Response({
            'total': len(result),
            'oblast_id': oblast.id,
            'oblast_name': oblast.name,
            'culture_id': culture.id,
            'culture_name': culture.name,
            'season': season,
            'applications': result
        })
    
    @action(detail=True, methods=['post'], url_path='cultures/(?P<culture_id>[^/.]+)/trial-types/(?P<trial_type_id>[^/.]+)/add-participants')
    def add_participants(self, request, pk=None, culture_id=None, trial_type_id=None):
        """
        Добавить участников к типу испытания культуры в плане
        
        POST /api/trial-plans/{id}/cultures/{culture_id}/trial-types/{trial_type_id}/add-participants/
        {
            "participants": [
                {
                    "patents_sort_id": 1774,
                    "statistical_group": 0,
                    "seeds_provision": "provided",
                    "maturity_group": "D03",
                    "application": 22,
                    "trials": [
                        {
                            "region_id": 1,
                            "predecessor": "fallow",  // или culture_id (например: 5)
                            "seeding_rate": 4.5,
                            "season": "spring"  // опционально, подставляется из культуры
                        }
                    ]
                }
            ]
        }
        """
        trial_plan = self.get_object()
        
        # Проверяем что культура существует в плане
        try:
            trial_plan_culture = TrialPlanCulture.objects.get(
                trial_plan=trial_plan,
                culture_id=culture_id,
                is_deleted=False
            )
        except TrialPlanCulture.DoesNotExist:
            return Response(
                {'error': 'Культура не найдена в плане'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Проверяем что тип испытания существует для этой культуры
        try:
            culture_trial_type = TrialPlanCultureTrialType.objects.get(
                trial_plan_culture=trial_plan_culture,
                trial_type_id=trial_type_id,
                is_deleted=False
            )
        except TrialPlanCultureTrialType.DoesNotExist:
            return Response(
                {'error': 'Тип испытания не найден для данной культуры в плане'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = TrialPlanAddParticipantsSerializer(data=request.data)
        if serializer.is_valid():
            participants_data = serializer.validated_data['participants']
            
            # ✅ ГРУППИРОВКА: Объединяем участников с одинаковым сортом
            # Ключ: patents_sort_id, значение: объединенные данные участника
            grouped_participants = {}
            
            for p_data in participants_data:
                sort_id = p_data['patents_sort_id']
                trials_data = p_data.pop('trials', [])
                
                if sort_id not in grouped_participants:
                    # Первый участник с этим сортом - сохраняем его данные
                    grouped_participants[sort_id] = p_data.copy()
                    grouped_participants[sort_id]['trials'] = trials_data
                else:
                    # Сорт уже есть - добавляем trials к существующему
                    grouped_participants[sort_id]['trials'].extend(trials_data)
            
            created_participants = []
            created_trials = []
            
            # Создать участников и их trials
            for sort_id, p_data in grouped_participants.items():
                # Получить trials data
                trials_data = p_data.pop('trials', [])
                
                # Получить application если указан
                application_id = p_data.get('application')
                application_obj = None
                if application_id:
                    try:
                        from .models import Application
                        application_obj = Application.objects.get(id=application_id)
                    except Application.DoesNotExist:
                        return Response({
                            'error': f'Application with ID {application_id} not found'
                        }, status=400)
                
                # ✅ ПРОВЕРКА: Если участник с таким сортом уже существует, используем его
                existing_participant = TrialPlanParticipant.objects.filter(
                    culture_trial_type=culture_trial_type,
                    patents_sort_id=sort_id,
                    is_deleted=False
                ).first()
                
                if existing_participant:
                    # Используем существующего участника
                    participant = existing_participant
                else:
                    # Создаем нового участника
                    # Автоматически генерируем participant_number (следующий доступный номер)
                    max_number = TrialPlanParticipant.objects.filter(
                        culture_trial_type=culture_trial_type,
                        is_deleted=False
                    ).aggregate(max_num=django_models.Max('participant_number'))['max_num'] or 0
                    participant_number = max_number + 1
                    
                    try:
                        participant = TrialPlanParticipant.objects.create(
                            culture_trial_type=culture_trial_type,
                            patents_sort_id=p_data['patents_sort_id'],
                            statistical_group=p_data.get('statistical_group', 1),
                            seeds_provision=p_data.get('seeds_provision', 'not_provided'),
                            participant_number=participant_number,
                            maturity_group=p_data.get('maturity_group', ''),
                            application=application_obj,
                            created_by=request.user
                        )
                        created_participants.append(participant)
                    except Exception as e:
                        return Response({
                            'error': f'Failed to create participant: {str(e)}'
                        }, status=400)
                
                # Создать trials для этого участника
                for trial_data in trials_data:
                    # Валидация обязательных полей
                    region_id = trial_data.get('region_id')
                    if not region_id:
                        return Response({
                            'error': 'region_id is required for each trial'
                        }, status=400)
                    
                    predecessor = trial_data.get('predecessor')
                    if not predecessor:
                        return Response({
                            'error': 'predecessor is required for each trial'
                        }, status=400)
                    
                    seeding_rate = trial_data.get('seeding_rate')
                    if seeding_rate is None:
                        return Response({
                            'error': 'seeding_rate is required for each trial'
                        }, status=400)
                    
                    # Подставляем дефолты с плана если не переданы
                    trial_season = trial_data.get('season', culture_trial_type.season)
                    
                    try:
                        trial = TrialPlanTrial.objects.create(
                            participant=participant,
                            region_id=region_id,
                            predecessor=predecessor,
                            seeding_rate=seeding_rate,
                            season=trial_season,
                            created_by=request.user
                        )
                    except Exception as e:
                        return Response({
                            'error': f'Failed to create trial: {str(e)}'
                        }, status=400)
                    created_trials.append(trial)
            
            # Обновить статус плана
            trial_plan.status = 'structured'
            trial_plan.update_statistics()
            
            return Response({
                'success': True,
                'message': f'Added {len(created_participants)} participants with {len(created_trials)} trials',
                'participants_created': len(created_participants),
                'trials_created': len(created_trials),
                'plan': TrialPlanSerializer(trial_plan).data
            })
        else:
            return Response(serializer.errors, status=400)
    
    @action(detail=True, methods=['post'], url_path='distribute')
    def distribute(self, request, pk=None):
        """
        Распределить план - создать PlannedDistribution для каждой заявки
        
        Использует trial-level season и trial_type из TrialPlanTrial.
        
        POST /api/v1/trial-plans/{id}/distribute/
        """
        trial_plan = self.get_object()
        
        if trial_plan.status not in ['structured']:
            return Response({
                'error': 'Plan must be structured before distribution'
            }, status=400)
        
        # Работаем с моделями, а не с JSON
        participants = TrialPlanParticipant.objects.filter(
            trial_plan=trial_plan,
            is_deleted=False
        ).prefetch_related('trials')
        
        created_distributions = []
        
        # Найти участников с заявками
        for participant in participants:
            if not participant.application:
                continue
            
            # Получить trials для участника
            trials = participant.trials.filter(is_deleted=False)
            
            # Создать PlannedDistribution для каждого trial
            for trial in trials:
                try:
                    # Создать или обновить PlannedDistribution
                    # Используем trial-level значения (не с плана!)
                    planned_dist, created = PlannedDistribution.objects.update_or_create(
                        application=participant.application,
                        region=trial.region,
                        defaults={
                            'trial_type': trial.trial_type,  # Из trial, не из плана
                            'planting_season': trial.season,  # Из trial, не из плана
                            'created_by': request.user,
                            'notes': f'Created from plan {trial_plan.id}'
                        }
                    )
                    
                    if created:
                        created_distributions.append({
                            'application_id': participant.application.id,
                            'application_number': participant.application.application_number,
                            'region_id': trial.region.id,
                            'region_name': trial.region.name,
                            'planned_distribution_id': planned_dist.id,
                            'trial_type': trial.trial_type.name if trial.trial_type else None,
                            'season': trial.season
                        })
                
                except Exception as e:
                    # Логируем ошибку и продолжаем
                    print(f"Error creating PlannedDistribution: {str(e)}")
                    continue
        
        # Обновить статус плана
        trial_plan.status = 'distributed'
        trial_plan.save()
        
        return Response({
            'success': True,
            'message': f'Plan distributed with {len(created_distributions)} new distributions',
            'created_distributions': created_distributions,
            'plan': TrialPlanSerializer(trial_plan).data
        })
    
    @action(detail=True, methods=['get'], url_path='statistics')
    def statistics(self, request, pk=None):
        """
        Получить статистику плана
        
        Считает по модели с учетом trial-level значений.
        
        GET /api/v1/trial-plans/{id}/statistics/
        """
        trial_plan = self.get_object()
        
        # Получить участников из модели (не JSON)
        participants = TrialPlanParticipant.objects.filter(
            trial_plan=trial_plan,
            is_deleted=False
        )
        
        # Статистика по группам спелости
        maturity_groups = {}
        for participant in participants:
            group = participant.maturity_group or 'unknown'
            if group not in maturity_groups:
                maturity_groups[group] = 0
            maturity_groups[group] += 1
        
        # Статистика по обеспеченности семенами
        seeds_provision = {}
        for participant in participants:
            provision = participant.seeds_provision or 'unknown'
            if provision not in seeds_provision:
                seeds_provision[provision] = 0
            seeds_provision[provision] += 1
        
        # Статистика по источникам (заявки vs реестр)
        from_applications = participants.filter(application__isnull=False).count()
        from_registry = participants.filter(application__isnull=True).count()
        
        # Статистика по типам испытания (из TrialPlanTrial)
        trial_types_stats = {}
        trials = TrialPlanTrial.objects.filter(
            participant__trial_plan=trial_plan,
            is_deleted=False
        ).select_related('trial_type')
        
        for trial in trials:
            trial_type_name = trial.trial_type.name if trial.trial_type else 'Не указан'
            if trial_type_name not in trial_types_stats:
                trial_types_stats[trial_type_name] = 0
            trial_types_stats[trial_type_name] += 1
        
        # Статистика по сезонам (из TrialPlanTrial)
        seasons_stats = {}
        for trial in trials:
            season = trial.season or 'unknown'
            if season not in seasons_stats:
                seasons_stats[season] = 0
            seasons_stats[season] += 1
        
        return Response({
            'plan_id': trial_plan.id,
            'total_participants': participants.count(),
            'from_applications': from_applications,
            'from_registry': from_registry,
            'maturity_groups': maturity_groups,
            'seeds_provision': seeds_provision,
            'total_trials': trial_plan.get_trials_count(),
            'trial_types': trial_types_stats,
            'seasons': seasons_stats,
            'status': trial_plan.status
        })
    
    @action(detail=False, methods=['get'], url_path='by-year')
    def by_year(self, request):
        """
        Получить планы по году
        
        GET /api/v1/trial-plans/by-year/?year=2025
        """
        year = request.query_params.get('year')
        if not year:
            return Response({'error': 'year parameter is required'}, status=400)
        
        try:
            year = int(year)
        except ValueError:
            return Response({'error': 'year must be a number'}, status=400)
        
        plans = self.get_queryset().filter(year=year)
        serializer = self.get_serializer(plans, many=True)
        
        return Response({
            'year': year,
            'total_plans': plans.count(),
            'plans': serializer.data
        })
    
    @action(detail=False, methods=['get'], url_path='predecessor-options')
    def predecessor_options(self, request):
        """
        Получить список доступных опций для предшественника
        
        GET /api/v1/trial-plans/predecessor-options/
        
        Returns:
        {
            "fallow": "Fallow (Пар)",
            "cultures": [
                {"id": 1, "name": "Wheat", "patents_id": 100},
                {"id": 2, "name": "Barley", "patents_id": 101}
            ]
        }
        """
        # Получаем список культур из локальной базы
        from trials_app.models import Culture
        cultures = Culture.objects.filter(is_deleted=False).order_by('name')
        
        culture_options = []
        for culture in cultures:
            culture_options.append({
                'id': culture.id,
                'name': culture.name,
                'patents_id': culture.culture_id
            })
        
        return Response({
            'fallow': 'Fallow (Пар)',
            'cultures': culture_options
        })
    
    @action(detail=True, methods=['get'], url_path='cultures')
    def get_cultures(self, request, pk=None):
        """
        Получить список культур в плане
        
        GET /api/v1/trial-plans/{id}/cultures/
        
        Returns:
        [
            {
                "id": 1,
                "culture": 5,
                "culture_name": "Пшеница",
                "culture_group": "Зерновые",
                "created_by": 1,
                "created_at": "2024-01-01T00:00:00Z",
                "updated_at": "2024-01-01T00:00:00Z"
            }
        ]
        """
        trial_plan = self.get_object()
        cultures = TrialPlanCulture.objects.filter(
            trial_plan=trial_plan, 
            is_deleted=False
        ).order_by('culture__name')
        
        serializer = TrialPlanCultureSerializer(cultures, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='add-culture')
    def add_culture(self, request, pk=None):
        """
        Добавить культуру в план
        
        POST /api/v1/trial-plans/{id}/add-culture/
        
        Body:
        {
            "culture_id": 5
        }
        
        Returns:
        {
            "id": 1,
            "culture": 5,
            "culture_name": "Пшеница",
            "culture_group": "Зерновые",
            "created_by": 1,
            "created_at": "2024-01-01T00:00:00Z",
            "updated_at": "2024-01-01T00:00:00Z"
        }
        """
        trial_plan = self.get_object()
        serializer = TrialPlanAddCultureSerializer(data=request.data)
        
        if serializer.is_valid():
            culture_id = serializer.validated_data['culture_id']
            
            # Проверяем, не добавлена ли уже эта культура
            existing = TrialPlanCulture.objects.filter(
                trial_plan=trial_plan,
                culture_id=culture_id,
                is_deleted=False
            ).exists()
            
            if existing:
                return Response(
                    {'error': 'Культура уже добавлена в план'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Создаем связь
            trial_plan_culture = TrialPlanCulture.objects.create(
                trial_plan=trial_plan,
                culture_id=culture_id,
                created_by=request.user
            )
            
            response_serializer = TrialPlanCultureSerializer(trial_plan_culture)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['delete'], url_path='remove-culture/(?P<culture_id>[^/.]+)')
    def remove_culture(self, request, pk=None, culture_id=None):
        """
        Удалить культуру из плана
        
        DELETE /api/v1/trial-plans/{id}/remove-culture/{culture_id}/
        
        Returns:
        {
            "message": "Культура удалена из плана"
        }
        """
        trial_plan = self.get_object()
        
        try:
            trial_plan_culture = TrialPlanCulture.objects.get(
                trial_plan=trial_plan,
                culture_id=culture_id,
                is_deleted=False
            )
            trial_plan_culture.delete()  # Soft delete
            return Response({'message': 'Культура удалена из плана'})
        
        except TrialPlanCulture.DoesNotExist:
            return Response(
                {'error': 'Культура не найдена в плане'}, 
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'], url_path='cultures/(?P<culture_id>[^/.]+)/add-trial-type')
    def add_trial_type_to_culture(self, request, pk=None, culture_id=None):
        """
        Добавить тип испытания к культуре в плане
        
        POST /api/trial-plans/{id}/cultures/{culture_id}/add-trial-type/
        
        Body:
        {
            "trial_type_id": 1,
            "season": "spring"
        }
        
        Returns:
        {
            "id": 1,
            "trial_type_id": 1,
            "trial_type_name": "КСИ",
            "season": "spring",
            "participants": [],
            "created_by": 1,
            "created_at": "2024-01-01T00:00:00Z",
            "updated_at": "2024-01-01T00:00:00Z"
        }
        """
        trial_plan = self.get_object()
        
        # Проверяем что культура существует в плане
        try:
            trial_plan_culture = TrialPlanCulture.objects.get(
                trial_plan=trial_plan,
                culture_id=culture_id,
                is_deleted=False
            )
        except TrialPlanCulture.DoesNotExist:
            return Response(
                {'error': 'Культура не найдена в плане'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        trial_type_id = request.data.get('trial_type_id')
        season = request.data.get('season', 'spring')
        
        if not trial_type_id:
            return Response(
                {'error': 'trial_type_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверяем, не добавлен ли уже этот тип испытания
        existing = TrialPlanCultureTrialType.objects.filter(
            trial_plan_culture=trial_plan_culture,
            trial_type_id=trial_type_id,
            is_deleted=False
        ).exists()
        
        if existing:
            return Response(
                {'error': 'Этот тип испытания уже добавлен к культуре'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверяем что trial_type существует
        try:
            trial_type = TrialType.objects.get(id=trial_type_id, is_deleted=False)
        except TrialType.DoesNotExist:
            return Response(
                {'error': 'Тип испытания не найден'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Создаем связь
        culture_trial_type = TrialPlanCultureTrialType.objects.create(
            trial_plan_culture=trial_plan_culture,
            trial_type=trial_type,
            season=season,
            created_by=request.user
        )
        
        # Формируем ответ
        response_data = {
            'id': culture_trial_type.id,
            'trial_type_id': trial_type.id,
            'trial_type_name': trial_type.name,
            'season': culture_trial_type.season,
            'participants': [],
            'created_by': request.user.id,
            'created_at': culture_trial_type.created_at,
            'updated_at': culture_trial_type.updated_at
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'], url_path='cultures/(?P<culture_id>[^/.]+)/trial-types/(?P<trial_type_id>[^/.]+)/add-participant')
    def add_participant_to_trial_type(self, request, pk=None, culture_id=None, trial_type_id=None):
        """
        Добавить участника к типу испытания культуры в плане
        
        POST /api/trial-plans/{id}/cultures/{culture_id}/trial-types/{trial_type_id}/add-participant/
        
        Body:
        {
            "patents_sort_id": 1774,
            "statistical_group": 0,
            "seeds_provision": "provided",
            "maturity_group": "D03",
            "application_id": 22
        }
        
        Returns:
        {
            "id": 1,
            "patents_sort_id": 1774,
            "statistical_group": 0,
            "seeds_provision": "provided",
            "maturity_group": "D03",
            "participant_number": 1,
            "application_id": 22,
            "created_by": 1,
            "created_at": "2024-01-01T00:00:00Z"
        }
        """
        trial_plan = self.get_object()
        
        # Проверяем что культура существует в плане
        try:
            trial_plan_culture = TrialPlanCulture.objects.get(
                trial_plan=trial_plan,
                culture_id=culture_id,
                is_deleted=False
            )
        except TrialPlanCulture.DoesNotExist:
            return Response(
                {'error': 'Культура не найдена в плане'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Проверяем что тип испытания существует для этой культуры
        try:
            culture_trial_type = TrialPlanCultureTrialType.objects.get(
                trial_plan_culture=trial_plan_culture,
                trial_type_id=trial_type_id,
                is_deleted=False
            )
        except TrialPlanCultureTrialType.DoesNotExist:
            return Response(
                {'error': 'Тип испытания не найден для данной культуры в плане'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Валидируем данные участника
        required_fields = ['patents_sort_id']
        for field in required_fields:
            if field not in request.data:
                return Response(
                    {'error': f'Поле {field} обязательно'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        patents_sort_id = request.data['patents_sort_id']
        
        # Проверяем, не добавлен ли уже этот сорт как участник
        existing_participant = TrialPlanParticipant.objects.filter(
            culture_trial_type=culture_trial_type,
            patents_sort_id=patents_sort_id,
            is_deleted=False
        ).first()
        
        if existing_participant:
            return Response(
                {'error': 'Этот сорт уже добавлен как участник'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Автоматически генерируем participant_number
        max_number = TrialPlanParticipant.objects.filter(
            culture_trial_type=culture_trial_type,
            is_deleted=False
        ).aggregate(max_num=django_models.Max('participant_number'))['max_num'] or 0
        participant_number = max_number + 1
        
        # Получаем application если указан
        application_id = request.data.get('application_id')
        application_obj = None
        if application_id:
            try:
                from .models import Application
                application_obj = Application.objects.get(id=application_id)
            except Application.DoesNotExist:
                return Response(
                    {'error': f'Application with ID {application_id} not found'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Создаем участника
        try:
            participant = TrialPlanParticipant.objects.create(
                culture_trial_type=culture_trial_type,
                patents_sort_id=patents_sort_id,
                statistical_group=request.data.get('statistical_group', 1),
                seeds_provision=request.data.get('seeds_provision', 'not_provided'),
                participant_number=participant_number,
                maturity_group=request.data.get('maturity_group', ''),
                application=application_obj,
                created_by=request.user
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to create participant: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Формируем ответ
        response_data = {
            'id': participant.id,
            'patents_sort_id': participant.patents_sort_id,
            'statistical_group': participant.statistical_group,
            'seeds_provision': participant.seeds_provision,
            'maturity_group': participant.maturity_group,
            'participant_number': participant.participant_number,
            'application_id': participant.application.id if participant.application else None,
            'created_by': request.user.id,
            'created_at': participant.created_at,
            'updated_at': participant.updated_at
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'], url_path='create-trial')
    def create_trial(self, request, pk=None):
        """
        Создать Trial из плана для конкретного ГСУ и культуры
        
        POST /api/v1/trial-plans/{id}/create-trial/
        {
            "region_id": 1,
            "culture_id": 5,
            "start_date": "2025-05-15",
            "responsible_person": "Иванов И.И.",
            "harvest_timing": "medium_early"  // опционально
        }
        
        Response:
        {
            "success": true,
            "trial_id": 20,
            "participants_count": 11,
            "status": "active",
            "message": "Trial created with 11 participants from plan"
        }
        """
        trial_plan = self.get_object()
        
        region_id = request.data.get('region_id')
        culture_id = request.data.get('culture_id')
        start_date = request.data.get('start_date')
        responsible_person = request.data.get('responsible_person')
        harvest_timing = request.data.get('harvest_timing')
        
        if not region_id or not culture_id or not start_date:
            return Response({
                'error': 'region_id, culture_id and start_date are required'
            }, status=400)
        
        # Проверить регион и культуру
        try:
            region = Region.objects.get(id=region_id, is_deleted=False)
            culture = Culture.objects.get(id=culture_id, is_deleted=False)
        except (Region.DoesNotExist, Culture.DoesNotExist):
            return Response({'error': 'Region or Culture not found'}, status=404)
        
        # Проверить что культура есть в плане
        culture_in_plan = TrialPlanCulture.objects.filter(
            trial_plan=trial_plan,
            culture=culture,
            is_deleted=False
        ).exists()
        
        if not culture_in_plan:
            return Response({
                'error': f'Culture {culture.name} not found in this plan'
            }, status=400)
        
        # Найти участников плана для этого ГСУ и культуры
        # Через TrialPlanTrial находим всех участников
        trial_plan_trials = TrialPlanTrial.objects.filter(
            participant__trial_plan=trial_plan,
            region=region,
            is_deleted=False
        ).select_related('participant').prefetch_related('participant__trials')
        
        if not trial_plan_trials.exists():
            return Response({
                'error': f'No participants found in plan for region {region.name} and culture {culture.name}'
            }, status=400)
        
        # Получить уникальных участников
        participants_to_add = []
        seen_sort_ids = set()
        
        for plan_trial in trial_plan_trials:
            participant = plan_trial.participant
            
            # Проверяем что сорт соответствует нужной культуре
            # (получаем сорт из Patents и проверяем его культуру)
            sort_record, _ = SortRecord.objects.get_or_create(
                sort_id=participant.patents_sort_id,
                defaults={'name': f'Сорт {participant.patents_sort_id}'}
            )
            
            # Синхронизировать если нужно
            if not sort_record.culture:
                sort_record.sync_from_patents()
            
            # Проверить что культура совпадает
            if sort_record.culture and sort_record.culture.id != culture.id:
                continue
            
            # Избегаем дубликатов
            if participant.patents_sort_id in seen_sort_ids:
                continue
            
            seen_sort_ids.add(participant.patents_sort_id)
            participants_to_add.append({
                'participant': participant,
                'sort_record': sort_record,
                'plan_trial': plan_trial
            })
        
        if not participants_to_add:
            return Response({
                'error': f'No matching participants found for culture {culture.name}'
            }, status=400)
        
        # Взять параметры из первого TrialPlanTrial
        first_plan_trial = participants_to_add[0]['plan_trial']
        
        # Создать Trial
        trial = Trial.objects.create(
            region=region,
            culture=culture,
            trial_type=first_plan_trial.trial_type or trial_plan.trial_type,
            start_date=start_date,
            year=trial_plan.year,
            planting_season=first_plan_trial.season,
            predecessor_culture=self._get_predecessor_culture(first_plan_trial.predecessor),
            growing_conditions='rainfed',  # По умолчанию богара
            cultivation_technology='traditional',  # По умолчанию обычная
            harvest_timing=harvest_timing,
            responsible_person=responsible_person,
            trial_plan=trial_plan,
            trial_plan_source='application',
            status='active',
            created_by=request.user
        )
        
        # Автоматически назначить показатели по культуре
        if culture.group_culture:
            auto_indicators = Indicator.objects.filter(
                group_cultures=culture.group_culture,
                is_deleted=False
            ).distinct()
            trial.indicators.set(auto_indicators)
        
        # Создать участников
        created_participants = []
        for idx, p_info in enumerate(participants_to_add, start=1):
            participant = p_info['participant']
            sort_record = p_info['sort_record']
            
            trial_participant = TrialParticipant.objects.create(
                trial=trial,
                sort_record=sort_record,
                statistical_group=participant.statistical_group,
                participant_number=idx,
                application_id=participant.application.id if participant.application else None
            )
            created_participants.append(trial_participant)
            
            # Автоматически создать пустые результаты для основных показателей
            create_basic_trial_results(trial_participant, request.user)
        
        # Обновить статус заявок
        application_ids = set()
        for tp in created_participants:
            if tp.application_id:
                application_ids.add(tp.application_id)
        
        if application_ids:
            Application.objects.filter(
                id__in=application_ids,
                status='distributed'
            ).update(status='in_progress')
        
        return Response({
            'success': True,
            'trial_id': trial.id,
            'participants_count': len(created_participants),
            'status': trial.status,
            'message': f'Trial created with {len(created_participants)} participants from plan',
            'trial': TrialSerializer(trial).data
        })
    
    def _get_predecessor_culture(self, predecessor_value):
        """
        Получить Culture объект предшественника
        
        Args:
            predecessor_value: 'fallow' или ID культуры
        
        Returns:
            Culture object или None
        """
        if predecessor_value == 'fallow':
            return None
        
        if isinstance(predecessor_value, int):
            try:
                return Culture.objects.get(id=predecessor_value, is_deleted=False)
            except Culture.DoesNotExist:
                return None
        
        return None
    
    @action(detail=False, methods=['get'], url_path='my-tasks')
    def my_tasks(self, request):
        """
        Получить задачи для агронома/филиала
        
        GET /api/v1/trial-plans/my-tasks/?region_id=1&year=2025
        
        Показывает какие Trial нужно создать из плана для конкретного ГСУ
        """
        region_id = request.query_params.get('region_id')
        year = request.query_params.get('year')
        
        if not region_id:
            return Response({
                'error': 'region_id is required'
            }, status=400)
        
        try:
            region = Region.objects.get(id=region_id, is_deleted=False)
        except Region.DoesNotExist:
            return Response({'error': 'Region not found'}, status=404)
        
        # Найти планы для области региона
        plans_query = TrialPlan.objects.filter(
            oblast=region.oblast,
            is_deleted=False
        )
        
        if year:
            plans_query = plans_query.filter(year=year)
        
        tasks = []
        
        for plan in plans_query:
            # Найти культуры в плане
            plan_cultures = TrialPlanCulture.objects.filter(
                trial_plan=plan,
                is_deleted=False
            ).select_related('culture')
            
            for plan_culture in plan_cultures:
                culture = plan_culture.culture
                
                # Подсчитать участников для этого ГСУ и культуры
                participants_count = TrialPlanTrial.objects.filter(
                    participant__trial_plan=plan,
                    region=region,
                    is_deleted=False
                ).values('participant__patents_sort_id').distinct().count()
                
                if participants_count == 0:
                    continue
                
                # Проверить создан ли уже Trial
                existing_trial = Trial.objects.filter(
                    trial_plan=plan,
                    region=region,
                    culture=culture,
                    is_deleted=False
                ).first()
                
                task_data = {
                    'plan_id': plan.id,
                    'plan_year': plan.year,
                    'culture_id': culture.id,
                    'culture_name': culture.name,
                    'culture_group': culture.group_culture.name if culture.group_culture else None,
                    'participants_count': participants_count,
                    'trial_created': existing_trial is not None,
                    'trial_id': existing_trial.id if existing_trial else None,
                    'trial_status': existing_trial.status if existing_trial else None,
                    'can_start': existing_trial is None  # Можно начать если еще не создан
                }
                
                tasks.append(task_data)
        
        return Response({
            'region_id': region.id,
            'region_name': region.name,
            'oblast_id': region.oblast.id,
            'oblast_name': region.oblast.name,
            'year': year,
            'total_tasks': len(tasks),
            'tasks': tasks
        })


# ============================================================================
# ГОДОВЫЕ ТАБЛИЦЫ РЕШЕНИЙ
# ============================================================================

class AnnualDecisionTableViewSet(viewsets.ModelViewSet):
    """
    Управление годовыми таблицами решений
    
    Endpoints:
    - GET /annual-decision-tables/ - список таблиц
    - POST /annual-decision-tables/ - создать новую таблицу
    - GET /annual-decision-tables/{id}/ - детали таблицы
    - PATCH /annual-decision-tables/{id}/ - обновить таблицу
    - POST /annual-decision-tables/{id}/finalize/ - завершить таблицу
    - GET /annual-decision-tables/{id}/export-excel/ - экспорт в Excel
    - GET /annual-decision-tables/{id}/export-pdf/ - экспорт в PDF
    """
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Список таблиц с фильтрацией"""
        queryset = AnnualDecisionTable.objects.filter(is_deleted=False)
        
        # Фильтры
        oblast_id = self.request.query_params.get('oblast_id')
        year = self.request.query_params.get('year')
        status = self.request.query_params.get('status')
        culture_id = self.request.query_params.get('culture_id')
        
        if oblast_id:
            queryset = queryset.filter(oblast_id=oblast_id)
        if year:
            queryset = queryset.filter(year=year)
        if status:
            queryset = queryset.filter(status=status)
        if culture_id:
            queryset = queryset.filter(culture_id=culture_id)
        
        return queryset.select_related('oblast', 'culture', 'created_by', 'finalized_by')
    
    def get_serializer_class(self):
        """Выбор serializer в зависимости от действия"""
        if self.action == 'create':
            return AnnualDecisionTableCreateSerializer
        elif self.action == 'retrieve':
            return AnnualDecisionTableDetailSerializer
        return AnnualDecisionTableSerializer
    
    @action(detail=True, methods=['post'])
    def finalize(self, request, pk=None):
        """
        Завершить таблицу (финализировать)
        
        POST /annual-decision-tables/{id}/finalize/
        
        Проверяет что все решения приняты и блокирует таблицу для редактирования
        """
        table = self.get_object()
        
        # Проверить что все решения приняты
        if not table.is_all_decisions_made():
            return Response({
                'success': False,
                'error': 'Не все решения приняты',
                'details': {
                    'total': table.get_items_count(),
                    'decided': table.get_decisions_count(),
                    'pending': table.get_items_count() - table.get_decisions_count()
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Завершить таблицу
        from django.utils import timezone
        table.status = 'finalized'
        table.finalized_by = request.user
        table.finalized_date = timezone.now().date()
        table.save()
        
        return Response({
            'success': True,
            'message': 'Таблица успешно завершена',
            'table': AnnualDecisionTableDetailSerializer(table).data
        })
    
    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """
        Экспорт таблицы в Excel
        
        GET /annual-decision-tables/{id}/export-excel/
        
        Response: Excel файл
        """
        table = self.get_object()
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from django.http import HttpResponse
            
            # Создать workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Решения {table.year}"
            
            # Заголовок
            culture_str = f" - {table.culture.name}" if table.culture else ""
            ws['A1'] = f"УРОЖАЙ КЛУБНЕЙ И ДРУГИЕ ПОКАЗАТЕЛИ ИСПЫТЫВАЕМЫХ СОРТОВ"
            ws['A2'] = f"{table.oblast.name}{culture_str}"
            ws['A3'] = f"{table.year} год"
            
            # Форматирование заголовка
            for row in [1, 2, 3]:
                ws[f'A{row}'].font = Font(bold=True, size=14)
                ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            # Заголовки колонок (строка 5)
            headers = [
                '№ п/п',
                'Сорта',
                'Группа спелости',
                'Урожайность 2022 (ц/га)',
                'Урожайность 2023 (ц/га)',
                'Урожайность 2024 (ц/га)',
                'Средняя урожайность за 3 года',
                'Отклонение от стандарта',
                'Решение',
                'Обоснование',
            ]
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Данные
            items = table.items.filter(is_deleted=False).order_by('row_number')
            
            for item in items:
                row_idx = 6 + item.row_number - 1
                
                yields = item.yields_by_year or {}
                
                ws.cell(row=row_idx, column=1, value=item.row_number)
                ws.cell(row=row_idx, column=2, value=item.sort_record.name)
                ws.cell(row=row_idx, column=3, value=item.maturity_group or '')
                ws.cell(row=row_idx, column=4, value=yields.get('2022', '') or yields.get(2022, ''))
                ws.cell(row=row_idx, column=5, value=yields.get('2023', '') or yields.get(2023, ''))
                ws.cell(row=row_idx, column=6, value=yields.get('2024', '') or yields.get(2024, ''))
                ws.cell(row=row_idx, column=7, value=item.average_yield or '')
                ws.cell(row=row_idx, column=8, value=item.deviation_from_standard or 'ст')
                ws.cell(row=row_idx, column=9, value=item.get_decision_display())
                ws.cell(row=row_idx, column=10, value=item.decision_justification or '')
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Создать HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"Tablica_resheniy_{table.oblast.code}_{table.year}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except ImportError:
            return Response({
                'success': False,
                'error': 'openpyxl не установлен'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """
        Получить статистику таблицы
        
        GET /annual-decision-tables/{id}/statistics/
        """
        table = self.get_object()
        stats = table.get_statistics()
        
        return Response({
            'table_id': table.id,
            'year': table.year,
            'oblast': table.oblast.name,
            'statistics': stats,
            'progress_percentage': table.get_progress_percentage(),
            'is_complete': table.is_all_decisions_made(),
        })


class AnnualDecisionItemViewSet(viewsets.ModelViewSet):
    """
    Управление элементами таблицы (решения по сортам)
    
    Endpoints:
    - GET /annual-decision-items/ - список элементов
    - GET /annual-decision-items/{id}/ - детали элемента
    - PATCH /annual-decision-items/{id}/ - обновить решение
    - POST /annual-decision-items/{id}/make-decision/ - принять решение
    - DELETE /annual-decision-items/{id}/reset-decision/ - сбросить решение
    """
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Список элементов с фильтрацией"""
        queryset = AnnualDecisionItem.objects.filter(is_deleted=False)
        
        # Фильтры
        table_id = self.request.query_params.get('table_id')
        decision = self.request.query_params.get('decision')
        sort_id = self.request.query_params.get('sort_id')
        
        if table_id:
            queryset = queryset.filter(annual_table_id=table_id)
        if decision:
            queryset = queryset.filter(decision=decision)
        if sort_id:
            queryset = queryset.filter(sort_record_id=sort_id)
        
        return queryset.select_related(
            'annual_table',
            'sort_record',
            'decided_by'
        ).order_by('annual_table', 'row_number')
    
    def get_serializer_class(self):
        """Выбор serializer"""
        if self.action in ['update', 'partial_update', 'make_decision']:
            return AnnualDecisionItemUpdateSerializer
        elif self.action == 'retrieve':
            return AnnualDecisionItemDetailSerializer
        return AnnualDecisionItemSerializer
    
    def update(self, request, *args, **kwargs):
        """Обновление элемента - проверка что таблица не finalized"""
        instance = self.get_object()
        
        if instance.annual_table.status == 'finalized':
            return Response({
                'success': False,
                'error': 'Невозможно изменить решение. Таблица завершена и заблокирована.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return super().update(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'])
    def make_decision(self, request, pk=None):
        """
        Принять решение по сорту
        
        POST /annual-decision-items/{id}/make-decision/
        
        Body:
        {
            "decision": "approved",
            "decision_justification": "...",
            "decision_recommendations": "...",
            "recommended_zones": [...]
        }
        """
        item = self.get_object()
        
        # Проверить что таблица не finalized
        if item.annual_table.status == 'finalized':
            return Response({
                'success': False,
                'error': 'Таблица завершена. Редактирование невозможно.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Обновить решение
        serializer = AnnualDecisionItemUpdateSerializer(
            item,
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        # Вернуть обновленные данные
        return Response({
            'success': True,
            'message': 'Решение успешно сохранено',
            'item': AnnualDecisionItemDetailSerializer(item).data,
            'table_progress': item.annual_table.get_progress_percentage()
        })
    
    @action(detail=True, methods=['delete'])
    def reset_decision(self, request, pk=None):
        """
        Сбросить решение (вернуть в статус pending)
        
        DELETE /annual-decision-items/{id}/reset-decision/
        """
        item = self.get_object()
        
        # Проверить что таблица не finalized
        if item.annual_table.status == 'finalized':
            return Response({
                'success': False,
                'error': 'Таблица завершена. Редактирование невозможно.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Сбросить решение
        item.decision = 'pending'
        item.decision_justification = ''
        item.decision_recommendations = ''
        item.decision_date = None
        item.decided_by = None
        item.continue_reason = None
        item.continue_until_year = None
        item.removal_reason = None
        item.save()
        
        return Response({
            'success': True,
            'message': 'Решение сброшено',
            'item': AnnualDecisionItemSerializer(item).data
        })
    
    @action(detail=True, methods=['post'])
    def refresh_data(self, request, pk=None):
        """
        Обновить агрегированные данные из испытаний
        
        POST /annual-decision-items/{id}/refresh-data/
        
        Пересчитывает урожайность и показатели из Trial/TrialResult
        """
        item = self.get_object()
        
        # Проверить что таблица не finalized
        if item.annual_table.status == 'finalized':
            return Response({
                'success': False,
                'error': 'Таблица завершена. Обновление данных невозможно.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Пересчитать данные
        item.aggregate_trial_data()
        
        return Response({
            'success': True,
            'message': 'Данные обновлены',
            'item': AnnualDecisionItemDetailSerializer(item).data
        })
